from genericpath import isfile
from itertools import count
from operator import le
from re import T
from services.hospital.reports import Reports
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from openpyxl.styles import numbers
from openpyxl import Workbook
from  openpyxl.styles import Font,Alignment,Border,Side
from openpyxl import load_workbook,styles
import json
import copy
import os
import shutil
from django.conf import  settings
from okb2.models import MyUser
from celery.contrib import rdb
from docxtpl import DocxTemplate

from abc import  ABC
from datetime import datetime
from services.hospital.patient import PatientsData
from services.hospital.reports import *
from collections import OrderedDict
import services.hospital.reference_reports as rr
import numpy
# from services.hospital.reference_reports import get_list_otd_prof 
# from services.hospital.reference_reports import get_list_otd_prof

from okb2.models import V001,Ds,Ab_Obsh
from hospital.models import Manpy

border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))

class Create(Reports):
    def __init__(self,user,request):
        super().__init__(user,request)
        self._user_group_name = 'hospital_annual_reports_%s' % user
        # self.list_data = json.loads(self.request['list_data']) if self.request.get('list_data') != None else []
        self._path_shoblons = 'shoblons/hospital/forms'
        self._path_shoblons_oth = 'shoblons/hospital/oth'
        self._temp_dir = 'temp'
        self._border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))


    def create(self):
        self._sluchays = self.get_sluchays()
        self.old_sluchays = copy.deepcopy(self._sluchays)
        type_fun = self._request.get('type_report')
        
        #Допилить с фильтрами
        if self._request.get('filters',None) != None:
            filters = json.loads(self._request.get('filters'))
            if filter:
                self._sluchays = self.filter(filters, self._sluchays)
        
        self._filename = self._request.get('filename', None)
        self._filename = f'{self._filename}.xlsx' if self._filename != None and len(self._filename) > 0 else f'{type_fun}_{self._user.user.id}.xlsx'

        if type_fun == 'annual_13_1_1':
            self._annual_13_1_1()
        elif type_fun == 'annual_13_1_2':
            self._annual_13_1_2()
        elif type_fun == 'annual_13_1_3':
            self._annual_13_1_3()
        elif type_fun == 'annual_13_1_4':
            self._annual_13_1_4()
        elif type_fun == 'annual_13_1_5':
            self._annual_13_1_5()
        elif type_fun == 'annual_14_1_1':
            self._annual_14_1_1()
        elif type_fun == 'annual_14_1_2':
            self._annual_14_1_2()
        elif type_fun == 'annual_14_1_3':
            self._annual_14_1_3()
        elif type_fun == 'annual_14_1_4':
            self._annual_14_1_4()
        elif type_fun == 'annual_14_1_5':
            self._annual_14_1_5()
        elif type_fun == 'annual_14_2_1':
            self._annual_14_2_1()
        elif type_fun == 'annual_14_3_1':
            self._annual_14_3_1()
        elif type_fun == 'annual_14_3_2':
            self._annual_14_3_2()
        elif type_fun == 'annual_14_3_3':
            self._annual_14_3_3()
        elif type_fun == 'annual_14_3_4':
            self._annual_14_3_4()
        elif type_fun == 'annual_14_3_5':
            self._annual_14_3_5()
        elif type_fun == 'annual_14_3_6':
            self._annual_14_3_6()
        elif type_fun == 'annual_14_3_7':
            self._annual_14_3_7()
        elif type_fun == 'annual_14_3_8':
            self._annual_14_3_8()
        elif type_fun == 'annual_30_1_1':
            self._annual_30_1_1()
        elif type_fun == 'annual_30_2_1':
            self._annual_30_2_1()
        elif type_fun == 'annual_30_2_2':
            self._annual_30_2_2()
        elif type_fun == 'annual_30_2_3':
            self._annual_30_2_3()
        elif type_fun == 'annual_30_3_1':
            self._annual_30_3_1()
        elif type_fun == 'annual_16_1_1':
            self._annual_16_1_1()
        elif type_fun == 'annual_57_1_1':
            self._annual_57_1_1()
        elif type_fun == 'annual_57_1_2':
            self._annual_57_1_2()
        elif type_fun == 'annual_57_1_3':
            self._annual_57_1_3()
        elif type_fun == 'annual_57_1_4':
            self._annual_57_1_4()
        elif type_fun == 'annual_pr_1':
            self._annual_pr_1()
        elif type_fun == 'annual_pr_2':
            self._annual_pr_2()
        elif type_fun == 'annual_pr_3':
            self._annual_pr_3()
        elif type_fun == 'annual_pr_4':
            self._annual_pr_4()
        elif type_fun == 'annual_pr_5':
            self._annual_pr_5()
        elif type_fun == 'annual_pr_6':
            self._annual_pr_6()
        elif type_fun == 'annual_pr_7':
            self._annual_pr_7()
        elif type_fun == 'annual_pr_8':
            self._annual_pr_8()
        elif type_fun == 'annual_pr_9':
            self._annual_pr_9()
        elif type_fun == 'annual_pr_a':
            self._annual_pr_a()
        elif type_fun == 'annual_pr_b':
            self._annual_pr_b()
        elif type_fun == 'annual_pr_v':
            self._annual_pr_v()
        elif type_fun == 'annual_pr_g':
            self._annual_pr_g()
        elif type_fun == 'annual_pr_d':
            self._annual_pr_d()
        


    def _annual_13_1_1(self):
        # print(self._user_group_name)
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            doc = DocxTemplate(file_new)
            context = {'i_all':111,
                       'i_1':222,
                       'i_2':333}
            doc.render(context)
            doc.save(file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_3(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_4(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_1(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_3(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_2_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_2(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_6(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_7(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_8(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_1_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_2_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_2_2(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_2_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_3_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})


    def _annual_16_1_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_2(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_pr_1(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT,self._path_shoblons_oth,'annual_pr_1.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT,self._temp_dir,str(self._user.id),file_new])
            shutil.copy2(file_shoblon,file_new)
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self._date_1.strftime("%d.%m.%Y")} по {self._date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self._user.statistics_type.name).capitalize()
            filetrs = ['otd','prof_k','count_sluchay','prof_k_n','goc_ek','rez_umer','rez_umer_goc_ek','rez_umer_goc_ek_sr']
            temp = self.filetr_sluchays(filetrs,self._sluchays)
            # res_set = [[r[0],r[1],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0] for r in set(rez for rez in temp)]
            row = 9
            # res_set = [r for r in list(set(rez for rez in temp))]

            res_set = set()
            for t in temp:
                x = (t[0],t[1],None,None,None,None,None,None,None,None,None,None,None,None,None,None,None)
                res_set.add(x)
            
            res_set = [list(r) for r in res_set]
            for t in temp:
                for r in res_set:
                    if r[0] == t[0] and r[1] == t[1]:
                        r[2] = r[2] + t[2] if r[2] != None else t[2]
                        r[3] = r[3] + t[3] if r[3] != None else t[3]
                        r[5] = r[5] + t[4] if r[5] != None else t[4]
                        r[8] = r[8] + t[5] if r[8] != None else t[5]
                        r[11] = r[11] + t[6] if r[11] != None else t[6]
                        r[13] = r[13] + t[7] if r[13]!= None else t[7]
            
            for r in res_set:
                try:
                    r[4] = float("{0:.2f}".format(r[3] / r[2] ))
                except ZeroDivisionError:
                    r[4] = r[3]

            max_otd = max(len(r[0]) for r in res_set)
            max_prof_k = max(len(r[1]) for r in res_set)
            col_otd = str(sheet.cell(column=1, row=6)).split('.')[1][:-1]
            col_max_prof_k = str(sheet.cell(column=2, row=6)).split('.')[1][:-1]
            sheet.column_dimensions[col_otd[0]].width = max_otd + 10
            sheet.column_dimensions[col_max_prof_k[0]].width = max_prof_k + 10

            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self._border
            # wb.save(file_new)
            self._path()
            wb.save(self._dir + self._filename)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'download', 'text': self._dir + self._filename,'name':'Отчет о работе отделений (по ИБ)'})
    def _annual_pr_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons_oth, 'annual_pr_2_1.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT,self._temp_dir,str(self._user.id),file_new])
            shutil.copy2(file_shoblon,file_new)
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self._date_1.strftime("%d.%m.%Y")} по {self._date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self._user.statistics_type.name).capitalize()
            row = 10
            filetrs = ['otd', 'prof_k','count_oper','goc_ek','count_oper_all','goc_ek_oper']
            temp = self.filetr_sluchays(filetrs, self._sluchays)
            res_set = set()
            for t in temp:
                x = (
                t[0], t[1], None, None, None, None, None, None, None, None, None, None, None, None)
                res_set.add(x)

            res_set = [list(r) for r in res_set]
            for t in temp:
                for r in res_set:
                    if r[0] == t[0] and r[1] == t[1]:
                        r[2] = r[2] + t[2] if r[2] != None else t[2]
                        r[3] = r[3] + t[3] if r[3] != None else t[3]
                        r[4] = r[4] + t[4] if r[4] != None else t[4]
                        r[5] = r[5] + t[5] if r[5] != None else t[5]

            max_otd = max(len(r[0]) for r in res_set)
            max_prof_k = max(len(r[1]) for r in res_set)
            col_otd = str(sheet.cell(column=1, row=6)).split('.')[1][:-1]
            col_max_prof_k = str(sheet.cell(column=2, row=6)).split('.')[1][:-1]
            sheet.column_dimensions[col_otd[0]].width = max_otd + 10
            sheet.column_dimensions[col_max_prof_k[0]].width = max_prof_k + 10

            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row, column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self._border

            self._path()
            wb.save(self._dir + self._filename)

        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'download', 'text': self._dir + self._filename,'name':'Отчет о хирур.работе отделений'})
    def _annual_pr_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_6(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_7(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_8(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_9(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_a(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_b(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_v(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_g(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_d(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

class AnnualReportABC(ABC):
    def __init__(self,user,request):
        self.user = MyUser.objects.get(user=user)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
        self.date_1 = datetime.strptime(self.request.get('date_1'),'%Y-%m-%d').date()
        self.date_2 = datetime.strptime(self.request.get('date_2'), '%Y-%m-%d').date()
        self.path_shoblons = 'shoblons/hospital/forms'
        self.path_shoblons_oth = 'shoblons/hospital/oth'
        self.temp_dir = 'temp'
        self.border = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='thin', color='000000'))
    def create(self):
        pass
    def is_file(self,file,forms=False):
        if forms:
            file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, file])
        else:
            file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons_oth, file])
        if os.path.exists(file_shoblon):
            filename = self.request.get('filename', None)
            type_fun = self.request.get('type_report')
            filename = f'{filename}.xlsx' if filename != None and len(filename) > 0 else f'{type_fun}_{self.user.user.id}.xlsx'
            file_new = filename
            path = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self.user.user.id),''])
            if not os.path.isdir(path):
                os.mkdir(path)

            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self.user.user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            return file_new
        return None
    def path(self):
        if not os.path.isdir(settings.MEDIA_ROOT + '/temp/' + f'{str(self.user.user.id)}/'):
            os.mkdir(settings.MEDIA_ROOT + '/temp/' + f'{str(self.user.user.id)}/')
        return settings.MEDIA_ROOT + '/temp/' + f'{str(self.user.user.id)}/'



def insert_sheet_AN_13_1_4(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']

    n1 = dict(all_=len(kwargs['list_vb_a']),up_to_12=0,from_12_to_22=0,pregnancy_n1=0,pregnancy_n2=0,pregnancy_n3=0)
    n1_1 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n1_2 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n1_3 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n2 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3 = dict(all_=len(kwargs['list_dskz_O']), up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_1 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_2_1 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_2_2 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_2_3 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)


    for v in kwargs['list_vb_a']:
        n1['up_to_12'] += 1 if v.vb_a.srber and  v.vb_a.srber < 12 else 0
        n1['from_12_to_22'] += 1 if v.vb_a.srber and   12 < v.vb_a.srber < 22 else 0
        n1['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
        n1['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
        n1['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

        if v.vb_a.pria and v.vb_a.pria.kod == 1:
            n1_1['all_'] += 1
            n1_1['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
            n1_1['from_12_to_22'] += 1 if v.vb_a.srber and 12 < v.vb_a.srber < 22 else 0
            n1_1['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
            n1_1['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
            n1_1['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

        if v.vb_a.pria and v.vb_a.pria.kod == 2:
            n1_2['all_'] += 1
            n1_2['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
            n1_2['from_12_to_22'] += 1 if v.vb_a.srber and 12 < v.vb_a.srber < 22 else 0
            n1_2['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
            n1_2['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
            n1_2['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

    for d in kwargs['list_dskz_O']:
        n3['up_to_12'] += 1 if d.vb_a.srber and d.vb_a.srber < 12 else 0
        n3['from_12_to_22'] += 1 if d.vb_a.srber and 12 < d.vb_a.srber < 22 else 0
        n3['pregnancy_n1'] += 1 if d.vb_a.n_ber and d.vb_a.n_ber == 1 else 0
        n3['pregnancy_n2'] += 1 if d.vb_a.n_ber and d.vb_a.n_ber == 2 else 0
        n3['pregnancy_n3'] += 1 if d.vb_a.n_ber and d.vb_a.n_ber >= 3 else 0

    print(n1)
    print(n3)





    # list_dskz_O
    # list_vb_a

    sheet.cell(row=3, column=1).value = str(name).capitalize()
    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'



def insert_sheet_APR_1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = rr.get_list_otd_prof(kwargs['data'])
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row = 8
    font = styles.Font(size=14, name='Arial')
    #Всего пациентов
    sl_all = []
    for d in data:
        for b in range(1,18):
            sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        row+=1
        sheet.row_dimensions[row].height = 45
        sheet.cell(row=row, column=1).value = d[0][0]
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        #Список итог всех пациентов в отделении
        otd_prof_sl_all = []
        for prof in d[1]:
            sheet.row_dimensions[row].height = 45
            sheet.cell(row=row, column=2).value = prof[0]
            sheet.cell(row=row, column=2).font = font
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = get_rez_apr_1(prof[1])
            otd_prof_sl_all.append(rez)
            sl_all.append(rez)
            for v in range(len(rez)):
                sheet.cell(row=row, column=3+v).value = rez[v] if rez[v] != 0 else None
                sheet.cell(row=row, column=3+v).font = font
                sheet.cell(row=row, column=3+v).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            else:
                for b in range(2,18):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            if len(d[1]) > 1:
                row+=1
        else:
            if len(d[1]) > 1:
                sheet.row_dimensions[row].height = 20
                sheet.cell(row=row, column=1).value = 'итого по отделению'
                sheet.cell(row=row, column=1).font = font
                r = None
                for o in range(len(otd_prof_sl_all)):
                    if o == 0:
                        r = numpy.array(otd_prof_sl_all[o])
                    else:
                        r+= numpy.array(otd_prof_sl_all[o])
                rez = r.tolist()
                rez = get_rez_apr_1(rez,False)
                for r in range(len(rez)):
                    sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                    sheet.cell(row=row, column=3+r).font = font
                    sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
    else:
        row+=1
        sheet.row_dimensions[row].height = 20
        for b in range(1,18):
            sheet.cell(row=row, column=b).border = styles.Border(top=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row, column=1).value = 'ВСЕГО ПО СТАЦИОНАРУ'
        sheet.cell(row=row, column=1).font = font
        r = None
        for o in range(len(sl_all)):
            if o == 0:
                r = numpy.array(sl_all[o])
            else:
                r+= numpy.array(sl_all[o])
        rez = r.tolist()
        rez = get_rez_apr_1(rez,False)
        for r in range(len(rez)):
            sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
            sheet.cell(row=row, column=3+r).font = font
            sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)

def insert_sheet_APR_2(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = rr.get_list_otd_prof(kwargs['data'])
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row = 7
    font = styles.Font(size=12, name='Arial')
    #Всего пациентов
    sl_all = []
    # print(data)
    count_all_otd = 0
    for d in data:
        for b in range(1,15):
            sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        row+=1
        count_otd = 0
        for prof in d[1]:
            count_otd += len(prof[1])
        print(count_otd)
        count_all_otd +=count_otd

        sheet.row_dimensions[row].height = 45
        sheet.cell(row=row, column=1).value = d[0][0]
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        #Список итог всех пациентов в отделении
        otd_prof_sl_all = []

        for prof in d[1]:
            sheet.row_dimensions[row].height = 45
            sheet.cell(row=row, column=2).value = prof[0]
            sheet.cell(row=row, column=2).font = font
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = get_rez_apr_2(prof[1],count=count_otd)
            otd_prof_sl_all.append(rez)
            sl_all.append(rez)
            for v in range(len(rez)):
                if v != 12:
                    sheet.cell(row=row, column=3+v).value = rez[v] if rez[v] != 0 else None
                    sheet.cell(row=row, column=3+v).font = font
                    sheet.cell(row=row, column=3+v).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            else:
                for b in range(2,15):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            if len(d[1]) > 1:
                row+=1
        else:
            if len(d[1]) > 1:
                sheet.row_dimensions[row].height = 20
                sheet.cell(row=row, column=1).value = 'итого по отделению'
                sheet.cell(row=row, column=1).font = font
                r = None
                for o in range(len(otd_prof_sl_all)):
                    if o == 0:
                        r = numpy.array(otd_prof_sl_all[o])
                    else:
                        r+= numpy.array(otd_prof_sl_all[o])
                rez = r.tolist()
                print(count_otd)
                rez = get_rez_apr_2(rez,False,count=count_otd)
                for r in range(len(rez)):
                    if r != 12:
                        sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                        sheet.cell(row=row, column=3+r).font = font
                        sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
    else:
        row+=1
        sheet.row_dimensions[row].height = 20
        for b in range(1,15):
            sheet.cell(row=row, column=b).border = styles.Border(top=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row, column=1).value = 'ВСЕГО ПО СТАЦИОНАРУ'
        sheet.cell(row=row, column=1).font = font
        r = None
        for o in range(len(sl_all)):
            if o == 0:
                r = numpy.array(sl_all[o])
            else:
                r+= numpy.array(sl_all[o])
        rez = r.tolist()
        rez = get_rez_apr_2(rez,False,count_all_otd)
        for r in range(len(rez)):
            if r != 12:
                sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                sheet.cell(row=row, column=3+r).font = font
                sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        row+=1
        sheet.merge_cells(f"A{sheet.cell(row=row, column=1).row}:N{sheet.cell(row=row, column=1).row}")
        sheet.cell(row=row, column=1).value = f'Всего предоперац.к\дн пл.б-х: {int(rez[11])}'
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        sheet.cell(row=row, column=1).font = font
        sheet.row_dimensions[row].height = 30

def insert_sheet_VRA(**kwargs):
    sheet = kwargs['sheet'][0]
    sheet1 = kwargs['sheet'][1]
    sheet2 = kwargs['sheet'][2]
    sheet3 = kwargs['sheet'][3]

    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    vra = kwargs['vra']
    font = styles.Font(size=12, name='Arial')
     ### Операции
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = f'Хирург - {vra.naim} {vra.ini} ({vra.kod})'
    oper_set = set()
    data = []
    for d in kwargs['data']:
        opers = rr.get_opers(d)
        if opers:
            for o in opers:
                if o.kodx and o.kodx.kod == vra.kod:
                    if o.kod_op:
                        oper_set.add(o.kod_op.kod)
    for o in oper_set:
        data.append([[o],[0,0,0,0,0,0]])

    for op in range(len(data)):
        for d in kwargs['data']:
            opers = rr.get_opers(d)
            if opers:
                for o in opers:
                    if o.kodx and o.kodx.kod == vra.kod:
                        if o.kod_op and o.kod_op.kod == data[op][0][0]:
                            data[op][1][0] += 1
                            if d.sluchay.goc and d.sluchay.goc.id_tip == 1:
                                data[op][1][2] += 1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                                data[op][1][3] += 1
                            if o.kodan:
                                data[op][1][4] += 1
                            # if o.oslo.count()>0:
                            #     data[op][1][5] += 1
    count = sum([d[1][0] for d in data])
    row = 8
    all_temp = [x[1] for x in data]
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(6)]
    all_temp[1] = 100

    for d in data:
        row+=1
        v001 = V001.objects.filter(kod=d[0][0])
        if v001.count() > 0:
            sheet.row_dimensions[row].height = 35
            sheet.cell(row=row, column=1).value = f'{v001[0].kod } - {v001[0].naim}'
            sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            sheet.cell(row=row, column=1).font = font
            sheet.cell(row=row, column=1).border = border
            for v in range(len(d[1])):
                if v == 0:
                    sheet.cell(row=row, column=v+2).value = d[1][v]
                    try:
                        d[1][1] = float('{0:.2f}'.format(d[1][v]*100/count))
                    except ZeroDivisionError:
                        d[1][1] = 0
                else:
                    sheet.cell(row=row, column=v+2).value = d[1][v] if d[1][v] != 0 else None
                sheet.cell(row=row, column=v+2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=v+2).font = font
                sheet.cell(row=row, column=v+2).border = border
    else:
        row+=1
        sheet.row_dimensions[row].height = 35
        sheet.cell(row=row, column=1).value = 'ИТОГО'
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).border = border
        for a in range(len(all_temp)):
            sheet.cell(row=row, column=a+2).value = all_temp[a] if all_temp[a] != 0 else None
            sheet.cell(row=row, column=a+2).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row, column=a+2).font = font
            sheet.cell(row=row, column=a+2).border = border
    ### Ноз.группы
    data = []
    ds_set = set()
    for d in kwargs['data']:
        if d.le_vr.kod and d.le_vr.kod.kod == vra.kod:
            if d.sluchay.dskz:
                ds_set.add(d.sluchay.dskz.kod[:3])

    for o in ds_set:
        data.append([[o],[0,0,0,0,0,0,0,0,0,0,0,0]])

    for ds in range(len(data)):
        for d in kwargs['data']:
            if d.le_vr.kod and d.le_vr.kod.kod == vra.kod:
                if d.sluchay.dskz and d.sluchay.dskz.kod[:3] == data[ds][0][0]:
                    data[ds][1][0] += 1
                    try:
                        data[ds][1][2] += d.le_vr.otd
                    except TypeError:
                        pass
                    if d.sluchay.goc and d.sluchay.goc.id_tip == 1: 
                        data[ds][1][3] += 1
                    if d.sluchay.goc and d.sluchay.goc.id_tip == 3:
                        data[ds][1][5] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 101:
                        data[ds][1][7] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 102:
                        data[ds][1][8] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 103:
                        data[ds][1][9] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 104:
                        data[ds][1][10] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz in [105,106]:
                        data[ds][1][11] += 1
    row = 9
    all_temp = [x[1] for x in data]
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(12)]
    sheet1.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet1.cell(row=5, column=1).value = f'{vra.naim} {vra.ini} ({vra.kod})'
    for d in data:
        row+=1
        ds = Ds.objects.filter(kod=d[0][0])[0]
        sheet1.row_dimensions[row].height = 35
        sheet1.cell(row=row, column=1).value = f'{ds.kod} {ds.naim}'
        sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)

        try:
            d[1][1] = float('{0:.2f}'.format(d[1][0]*100/all_temp[0]))
        except ZeroDivisionError:
            pass
        try:
            d[1][2] = float('{0:.2f}'.format(d[1][2]/d[1][0]))
        except ZeroDivisionError:
            pass
        try:
            d[1][4] = float('{0:.2f}'.format(d[1][3]*100/d[1][0]))
        except ZeroDivisionError:
            pass
        try:
            d[1][6] = float('{0:.2f}'.format(d[1][5]*100/d[1][0]))
        except ZeroDivisionError:
            pass

        for v in range(len(d[1])):
            sheet1.cell(row=row, column=2+v).value = d[1][v] if d[1][v] != 0 else None
            sheet1.cell(row=row, column=2+v).alignment = styles.Alignment(horizontal="center", vertical="center")
    else:
        row+=1
        sheet1.row_dimensions[row].height = 20
        sheet1.cell(row=row, column=1).value = 'ИТОГО'
        sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        all_temp[1] = 100

        try:
            all_temp[2] = float('{0:.2f}'.format(all_temp[2]/all_temp[0]))
        except ZeroDivisionError:
            pass
        try:
            all_temp[4] = float('{0:.2f}'.format(all_temp[3]*100/all_temp[0]))
        except ZeroDivisionError:
            pass
        try:
            all_temp[6] = float('{0:.2f}'.format(all_temp[5]*100/all_temp[0]))
        except ZeroDivisionError:
            pass

        for a in range(len(all_temp)):
            sheet1.cell(row=row, column=2+a).value = all_temp[a] if all_temp[a] != 0 else None
            sheet1.cell(row=row, column=2+a).alignment = styles.Alignment(horizontal="center", vertical="center")
    ###ДС
    ds_set = set()
    temp_data = []
    for d in kwargs['data']:
        if d.sluchay.manpy.count() > 0:
            mans = [Manpy.objects.get(id=m['id']) for m in d.sluchay.manpy.values('id')]
            for m in mans:
                if m.tnvr and m.tnvr.kod == vra.kod:
                    if m.kodmn and m.kodmn.kod == '01013':
                        temp_data.append(d)
                        if d.sluchay.dskz:
                            ds_set.add(d.sluchay.dskz.kod)
    data = []
    for o in ds_set:
        data.append([o,0])

    for ds in range(len(data)):
        for d in kwargs['data']:
            if d.sluchay.manpy.count() > 0:
                mans = [Manpy.objects.get(id=m['id']) for m in d.sluchay.manpy.values('id')]
                for m in mans:
                    if m.tnvr and m.tnvr.kod == vra.kod:
                        if m.kodmn and m.kodmn.kod == '01013':
                            if d.sluchay.dskz and d.sluchay.dskz.kod == data[ds][0]:
                                data[ds][1]+= 1
    row = 6

    sheet2.cell(row=2,column=1).value = f'Отчет за период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet2.cell(row=4, column=1).value = f'{vra.naim} {vra.ini} ({vra.kod})'

    all_count = sum([d[1] for d in data])
    for d in data:
        row+=1
        ds = Ds.objects.filter(kod=d[0])[0]
        sheet2.cell(row=row, column=1).value = f'{ds.kod} {ds.naim}'
        sheet2.cell(row=row, column=2).value = d[1]
        sheet2.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
    else:
        row+=1
        sheet2.cell(row=row, column=1).value = 'ИТОГО'
        sheet2.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet2.cell(row=row, column=2).value = all_count
        sheet2.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")

    ### man
    manpys = []
    kod_set = set()
    data = []
    row = 4
    for d in kwargs['data']:
        if d.sluchay.manpy.count() > 0:
            mans = [Manpy.objects.get(id=m['id']) for m in d.sluchay.manpy.values('id')]
            for m in mans:
                if m.tnvr and m.tnvr.kod == vra.kod:
                    manpys.append(m)
                    if m.kodmn:
                        kod_set.add(m.kodmn.kod)
    for o in kod_set:
        data.append([o,0])
    for d in range(len(data)):
        for m in manpys:
            if m.kodmn and data[d][0] == m.kodmn.kod:
                data[d][1]+= 1
    all_count = sum([d[1] for d in data])
    for d in data:
        row+=1
        am = Ab_Obsh.objects.filter(kod=d[0])[0]
        sheet3.cell(row=row, column=1).value = f'{am.kod} {am.ima}'
        sheet3.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")    
        sheet3.cell(row=row, column=2).value = d[1]
        sheet3.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")    
    else:
        row+=1
        sheet3.cell(row=row, column=1).value = 'ИТОГО'
        sheet3.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet3.cell(row=row, column=2).value = all_count
        sheet3.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet3.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")  


def get_rez_apr_1(data,d=None):
    if d == None:
        bf = BetterFilter()
        sp = CountSluchaySpecification() ^ ProfKNSpecification() ^\
            GocEkSpecification() ^ GocEkNSpecification() ^\
            RezUmerSpecification() ^ RezUmerKdSpecification() ^ \
            RezUmerGocEkSpecification() ^ RezUmerGocEkSrSpecification() ^\
            RezUmerDetSpecification() ^ RezUmerKdDetSpecification()
        
        ## sp = 0-CountSluchaySpecification() ^ 1-ProfKNSpecification() ^ 2-new-сред
            #  3-GocEkSpecification() ^ 4-new-% ^  5-GocEkNSpecification() ^\
            #  6-RezUmerSpecification() ^ 7-new-% ^ 8-RezUmerKdSpecification() ^ \
            #  9-RezUmerGocEkSpecification() ^ 10-new-% ^ 11-RezUmerGocEkSrSpecification() ^\
            #  12-RezUmerDetSpecification() ^ 13-new-% ^ 14-RezUmerKdDetSpecification()

        all_temp = []
        for patient in data:
            for p in bf.filter(patient,sp):
                temp = bf.format_list(p)
                for t in range(len(temp)):
                    if temp[t] == 'None':
                        temp[t] = 0
                all_temp.append([int(i) for i in temp])

        all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(10)]

        try:
            all_temp.insert(2,float('{0:.2f}'.format(all_temp[1]/all_temp[0])))
        except ZeroDivisionError:
            all_temp.insert(2,0)

        try:
            all_temp.insert(4,float('{0:.2f}'.format(all_temp[3]*100/all_temp[0])))
        except ZeroDivisionError:
            all_temp.insert(4,0)
        try:
            all_temp[5]= float('{0:.2f}'.format(all_temp[5]/all_temp[3]))
        except ZeroDivisionError:
            all_temp[5] = 0


        try:
            all_temp.insert(7,float('{0:.2f}'.format(all_temp[6]*100/all_temp[0])))
        except ZeroDivisionError:
            all_temp.insert(7,0)
        try:
            all_temp[8]= float('{0:.2f}'.format(all_temp[8]/all_temp[6]))
        except ZeroDivisionError:
            all_temp[8] = 0


        try:
            all_temp.insert(10,float('{0:.2f}'.format(all_temp[9]*100/all_temp[0])))
        except ZeroDivisionError:
            all_temp.insert(10,0)
        try:
            all_temp[11]= float('{0:.2f}'.format(all_temp[11]/all_temp[9]))
        except ZeroDivisionError:
            all_temp[11] = 0


        try:
            all_temp.insert(13,float('{0:.2f}'.format(all_temp[12]*100/all_temp[0])))
        except ZeroDivisionError:
            all_temp.insert(13,0)
        try:
            all_temp[14]= float('{0:.2f}'.format(all_temp[14]/all_temp[12]))
        except ZeroDivisionError:
            all_temp[14] = 0

        return all_temp
    
    try:
        data[2] = float('{0:.2f}'.format(data[1]/data[0]))
    except ZeroDivisionError:
        data[2] = 0

    try:
        data[4] = float('{0:.2f}'.format(data[3]*100/data[0]))
    except ZeroDivisionError:
        data[4] = 0
    try:
        data[5]= float('{0:.2f}'.format(data[5]/data[3]))
    except ZeroDivisionError:
        data[5] = 0

    try:
        data[7] = float('{0:.2f}'.format(data[6]*100/data[0]))
    except ZeroDivisionError:
        data[7] = 0

    try:
        data[8]= float('{0:.2f}'.format(data[8]/data[6]))
    except ZeroDivisionError:
        data[8] = 0
    
    try:
        data[10] = float('{0:.2f}'.format(data[9]*100/data[0]))
    except ZeroDivisionError:
        data[10] = 0
    try:
        data[11]= float('{0:.2f}'.format(data[11]/data[9]))
    except ZeroDivisionError:
        data[11] = 0
    
    try:
        data[13] = float('{0:.2f}'.format(data[12]*100/data[0]))
    except ZeroDivisionError:
        data[13] = 0
    try:
        data[14]= float('{0:.2f}'.format(data[14]/data[12]))
    except ZeroDivisionError:
        data[14] = 0

    return data
def get_rez_apr_2(data,d=None,count=0):
    if d == None:
        bf = BetterFilter()
        sp = OperCountSpecification() ^ OperCountGocEkSpecification() ^ OperAllCountSpecification() ^ OperAllCountGocEkSpecification() ^\
            OperAllKdSpecification() ^ PredOperSpecification() ^ OsloCountAllSpecification()^ RezUmerOperSpecification() ^ PredOperKdSpecification()
        all_temp = []
        for patient in data:
            for p in bf.filter(patient,sp):
                temp = bf.format_list(p)
                for t in range(len(temp)):
                    if temp[t] == 'None':
                        temp[t] = 0
                all_temp.append([int(i) for i in temp])
        all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(9)]
    
        print(all_temp)

        # all_temp.insert(5,0)
        all_temp.insert(6,0)
        all_temp.insert(8,0)
        all_temp.insert(10,0)
        all_temp.insert(12,copy.deepcopy(all_temp[4]))
        print(all_temp)
        try:
            all_temp[4]= float('{0:.2f}'.format(all_temp[4]/all_temp[0]))
        except ZeroDivisionError:
            all_temp[4] = 0
        try:
            all_temp[5]= float('{0:.2f}'.format(all_temp[11]/(all_temp[0]-all_temp[1])))
        except ZeroDivisionError:
            all_temp[5] = 0

        # try:
        #     all_temp[6]= float('{0:.2f}'.format(all_temp[0]*100/all_temp[2]))
        # except ZeroDivisionError:
        #     all_temp[6] = 0
        if count != 0:
            try:
                all_temp[6]= float('{0:.2f}'.format(all_temp[0]*100/count))
            except ZeroDivisionError:
                all_temp[6] = 0
        else:
            all_temp[6] = 0
        try:
            all_temp[8]= float('{0:.2f}'.format(all_temp[7]*100/all_temp[0]))
        except ZeroDivisionError:
            all_temp[8] = 0
        try:
            all_temp[10]= float('{0:.2f}'.format(all_temp[9]*100/all_temp[0]))
        except ZeroDivisionError:
            all_temp[10] = 0
        
        # print(all_temp)

        return all_temp
    try:
        data[4]= float('{0:.2f}'.format(data[12]/data[0]))
    except ZeroDivisionError:
        data[4] = 0
    try:
        data[5]= float('{0:.2f}'.format(data[11]/(data[0]-data[1])))
    except ZeroDivisionError:
        data[5] = 0
    # try:
    #     data[6]= float('{0:.2f}'.format(data[0]*100/data[2]))
    # except ZeroDivisionError:
    #     data[6] = 0
    if count != 0:
        try:
            data[6]= float('{0:.2f}'.format(data[0]*100/count))
        except ZeroDivisionError:
            data[6] = 0
    else:
        data[6] = 0
    try:
        data[8]= float('{0:.2f}'.format(data[7]*100/data[0]))
    except ZeroDivisionError:
        data[8] = 0
    try:
        data[10]= float('{0:.2f}'.format(data[9]*100/data[0]))
    except ZeroDivisionError:
        data[10] = 0

    return data

class AnnualPr1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
    def create(self):
        file = self.is_file('annual_pr_1.xlsx')
        if file is not None:
            wb = load_workbook(file)
            sheet = wb.active
            patients = PatientsData(self.date_1,self.date_2,self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_APR_1(**dic)

            wb.save(self.path() + f'annual_pr_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download',
                                                                                'text': self.path() + f'annual_pr_1_{self.user.user.id}.xlsx',
                                                                                'name': 'Отчет о работе отделений (по ИБ)'})
        else:
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'text': 'Отсутствует шаблон - annual_pr_1.xlsx'})
            raise FileNotFoundError('Отсутствует шаблон - annual_pr_1.xlsx')

class AnnualPr2(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
    def create(self):
        file = self.is_file('annual_pr_2.xlsx')
        if file is not None:
            wb = load_workbook(file)
            sheet = wb.active
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_APR_2(**dic)
                        
            wb.save(self.path() + f'annual_pr_2_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                            {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                                    'text': self.path() + f'annual_pr_2_{self.user.user.id}.xlsx',
                                                                                    'name': 'Отчет о хирур.работе отделений'})

        else:
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'text': 'Отсутствует шаблон - annual_pr_2.xlsx'})
            raise FileNotFoundError('Отсутствует шаблон - annual_pr_2.xlsx')
#Нужны данные чтобы доделать
class Annual_13_1_4(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
    def create(self):
        file = self.is_file('annual_13_1_4.xlsx',forms=True)
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            list_dskz_O = []
            list_vb_a = []
            for p in patients.patients:
                if  p.vb_a != None and p.sluchay.dskz and (p.sluchay.dskz.kod[:3] in ['O02','O03,','O04','O05','O06']):
                    list_dskz_O.append(p)
                if p.vb_a != None and p.sluchay.dskz and (p.sluchay.dskz.kod[:3] in ['O04']):
                    list_vb_a.append(p)

            dic = dict([('sheet', sheet), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('list_dskz_O', list_dskz_O),
                        ('list_vb_a',list_vb_a)])
            insert_sheet_AN_13_1_4(**dic)
            wb.save(self.path() + f'group_an_13_4_{self.user.user.id}.xlsx')
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_group_data_annual', 'text': 'Отчет cфромирован'})
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'download_data_annual','name': 'Отчет о выбывших'})

class Annual_Vra(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('annual_vra.xlsx')
        if file:
            wb = load_workbook(file)
            sheet=wb.get_sheet_by_name('Операции')
            sheet1=wb.get_sheet_by_name('Ноз.групп')
            sheet2=wb.get_sheet_by_name('ДС')
            sheet3=wb.get_sheet_by_name('Манипул.')

            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            v = self.request.get('vra').split(' ')
            vra = Vra.objects.filter(kod=v[0],naim=v[1],dateend=None)[:1][0]

            dic = dict([('sheet', [sheet,sheet1,sheet2,sheet3]), ('name', self.user.statistics_type.name),('date_1', self.date_1),
                ('date_2', self.date_2),('vra',vra),('data',patients.patients)])
            insert_sheet_VRA(**dic)
            wb.save(self.path() + f'group_an_vra_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data_annual', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_an_vra_{self.user.user.id}.xlsx'})

def AnnualReport(user,request):
    print('AnnualReport',request)
    type_fun = request.get('type_report')
    if type_fun == 'annual_pr_1':
        report = AnnualPr1(user,request)
        report.create()
    elif type_fun == 'annual_pr_2':
        report = AnnualPr2(user,request)
        report.create()

    elif type_fun == 'annual_13_1_4':
        report = Annual_13_1_4(user,request)
        report.create()
    elif type_fun == 'annual_vra':
        report = Annual_Vra(user,request)
        report.create()