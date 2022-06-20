from django.http import response
from services.hospital.reports import Reports
from openpyxl.styles import numbers
from openpyxl import Workbook
from  openpyxl.styles import Font,Alignment,Border,Side
from collections import Counter
import json
import time
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
import os
from django.http import HttpResponse
class Create(Reports):
    def __init__(self,user,request):
        super().__init__(user,request)
        self._user_group_name = 'hospital_reports_mix_%s' % user
        self.list_data = json.loads(request['list_data']) if request.get('list_data') != None else []
        self.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
    def create(self,request):
        self.sluchays = self.get_sluchays()
        filters = json.loads(request.get('filters'))['filter']
        self._filename = request.get('filename',None)
        self._filename = f'{self._filename}.xlsx' if self._filename != None and len(self._filename) > 0 else  f'oth_all_mix_{self._user.user.id}.xlsx'
        if len(filters) > 0:
            self.sluchays = self.filter(filters,self.sluchays)
        self.create_file_report()

    def create_file_report(self):
            self.wb = Workbook()
            self.sheet = self.wb.active
            row = 1
            end_row = 0
            start_max = 0
            # text = 'Список выбывших из ГБУЗ ТО ОКБ № 2'
            # self.sh_title(self.sheet,text)
            self.header_d()
            for i,data in enumerate(self.list_data):
                self.sheet.cell(column=1 + i, row=1).value = self.header.get(data,data)
                self.sheet.cell(column=1 + i, row=1).border = self.border
                self.sheet.cell(column=1 + i, row=1).alignment = Alignment(horizontal="center", vertical="center")
                col = str(self.sheet.cell(column=1 + i, row=1)).split('.')[1][:-1]
                if len(col) == 2:
                    self.sheet.column_dimensions[col[0]].width = self.header.get(data+'_width')

                elif len(col) == 3:
                    self.sheet.column_dimensions[col[:2]].width = self.header.get(data + '_width')

            self.sheet.row_dimensions[1].height = 28

            for i,sluchays in enumerate(self.sluchays):

                max_row = []
                for l, data in enumerate(self.list_data):
                    if data in ['oper_date','oper_tm','oper_py','oper_kod_op','oper_goc','oper_kodx','oper_pop','oper_pr_osob',
                                'oper_k_mm','oper_kodxa','oper_kodxa1','oper_obz','oper_kodan']:
                        max_row.append(len(sluchays['oper']) if len(sluchays['oper']) != None else 0,)
                        break

                for l, data in enumerate(self.list_data):
                    if data in ['oslo_tnvr','oslo_date','oslo_kod_osl','oslo_xosl','oslo_posl','oslo_aosl']:
                        max_row.append(len(sluchays['oslo']) if sluchays['oslo'] != None else 0,)
                        break

                for l, data in enumerate(self.list_data):
                    if data in ['man_date','man_tnvr','man_kodmn','man_kol','man_pl']:
                        max_row.append(len(sluchays['manpy']) if sluchays['manpy'] != None else 0,)
                        break

                if end_row != 0:
                    if i != 1:
                        i = end_row - 1
                    else:
                        i += start_max - 1
                else:
                    start_max = max(max_row) if len(max_row)  != 0 else 0



                for l,data in enumerate(self.list_data):
                    self.sheet.cell(column=1 + l, row=i + 1 + row).alignment = Alignment(horizontal="center", vertical="center")
                    self.sheet.cell(column=1 + l, row=i + 1 + row).number_format = numbers.FORMAT_TEXT
                    if data == 'nib':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].nib
                    elif data == 'fam':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['patient'].fam
                    elif data == 'im':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].im
                    elif data == 'ot':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['patient'].ot
                    elif data == 'pol':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['patient'].pol.polname if sluchays['patient'].pol != None else ''
                    elif data == 'datp':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['sluchay'].datp.strftime('%d.%m.%Y') if sluchays['sluchay'].datp != None else ''
                    elif data == 'tm_otd':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['sluchay'].tm_otd
                    elif data == 'datv':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['sluchay'].datv.strftime('%d.%m.%Y') if sluchays['sluchay'].datv != None else ''
                    elif data == 'datr':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['patient'].datr.strftime('%d.%m.%Y') if sluchays['patient'].datr != None else ''
                    elif data == 'otd':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['sluchay'].otd.naim if sluchays['sluchay'].otd != None else ''
                    elif data == 'm_roj_in':
                        self.sheet.cell(column=1+l, row=i+1+row).value = sluchays['patient'].m_roj
                    elif data == 'adr_in':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['patient'].adr
                    elif data == 'rab':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['patient'].rab
                    elif data == 'prof':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['patient'].prof
                    elif data == 'r_n':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['patient'].r_n.naim if sluchays['patient'].r_n != None else ''
                    elif data == 'in_t':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['patient'].in_t.name if sluchays['patient'].in_t != None else ''
                    elif data == 'lpy':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].lpy.naim if sluchays['sluchay'].lpy != None else ''
                    elif data == 'npr_num':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].npr_num
                    elif data == 'npr_date':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].npr_date.strftime('%d.%m.%Y') if sluchays['sluchay'].npr_date != None else ''
                    elif data == 'alg':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].alg_display() if sluchays['sluchay'].alg != None else ''
                    elif data == 'goc':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].goc.tip_name if sluchays['sluchay'].goc != None else ''
                    elif data == 'prpg':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].prpg.naim if sluchays['sluchay'].prpg != None else ''
                    elif data == 'vrez':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].vrez.naim if sluchays['sluchay'].vrez != None else ''
                    elif data == 'p_per':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].p_per.naim if sluchays['sluchay'].p_per != None else ''

                    #Сведения о диагнозах

                    elif data == 'dsny':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].dsny.kod if sluchays['sluchay'].dsny != None else ''
                    elif data == 'ds_0':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].ds_0.kod if sluchays['sluchay'].ds_0 != None else ''
                    elif data == 'dsk':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].dsk.kod if sluchays['sluchay'].dsk != None else ''
                    elif data == 'dskz':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].dskz.kod if sluchays['sluchay'].dskz != None else ''
                    elif data == 'ds_osl':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].ds_osl.kod if sluchays['sluchay'].ds_osl != None else ''
                    elif data == 'dsc':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].dsc.kod if sluchays['sluchay'].dsc != None else ''
                    elif data == 'dson':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].dson.kod if sluchays['sluchay'].dson != None else ''
                    elif data == 'dat_otd':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].dat_otd
                    elif data == 'tm_otd_d':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].tm_otd
                    elif data == 'icx':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].icx.iz_name if sluchays['sluchay'].icx != None else ''
                    elif data == 'rslt':
                        self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['sluchay'].rslt.tip_name if sluchays['sluchay'].rslt != None else ''
                    #Койко-дни
                    if sluchays['le_vr'] != None:
                        if data == 'koy_N':
                            self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['le_vr'].kd
                        elif data == 'koy_aro':
                            self. sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['le_vr'].aro
                        elif data == 'koy_otd':
                            self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['le_vr'].otd
                        elif data == 'koy_prof_k':
                            self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['le_vr'].prof_k.k_prname if sluchays['le_vr'].prof_k != None else None
                        elif data == 'koy_kod':
                            self.sheet.cell(column=1 + l, row=i + 1+row).value = sluchays['le_vr'].kod.kod if sluchays['le_vr'].kod != None else None
                    if sluchays['oper'] != None:
                        for o, oper in enumerate(sluchays['oper']):
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).alignment = Alignment(horizontal="center",vertical="center")
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).number_format = numbers.FORMAT_TEXT
                            if data == 'oper_date':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.dato.strftime('%d.%m.%Y') if oper.dato != None else ''
                                    # print(oper.dato)
                            elif data == 'oper_tm':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.tm_o
                            elif data == 'oper_py':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.py.naim if oper.py != None else ''
                            elif data == 'oper_kod_op':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.kod_op.kod if oper.kod_op != None else ''
                            elif data == 'oper_goc':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.goc.id_tip if oper.goc != None else ''
                            elif data == 'oper_kodx':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.kodx.kod if oper.kodx != None else ''
                            elif data == 'oper_pop':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.pop
                                #?????oper_pr_osob
                            # elif data == 'oper_pr_osob':
                            #     for o, data in enumerate(oper):
                            #         sheet.cell(column=1 + l, row=i + 1 + o+row).value = data.pr_osob
                            elif data == 'oper_pr_osob':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = ''
                            elif data == 'oper_k_mm':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.k_mm
                            elif data == 'oper_kodxa':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.kodxa.kod if oper.kodxa != None else ''
                            elif data == 'oper_kodxa1':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.kodxa1.kod if oper.kodxa1 != None else ''
                            elif data == 'oper_obz':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.obz.kod if oper.obz != None else ''
                            elif data == 'oper_kodan':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oper.kodan.kod if oper.kodan != None else ''

                    if data == 'ksg_osn':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].ksg_osn.code_usl  if sluchays['sluchay'].ksg_osn != None else ''
                    elif data == 'oopkk':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].oopkk.kod  if sluchays['sluchay'].oopkk != None else ''
                    elif data == 'ksg_sop':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].ksg_sop.code_usl if sluchays['sluchay'].ksg_sop != None else ''
                    elif data == 'iddoc':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].iddokt.kod  if sluchays['sluchay'].iddokt != None else ''
                    if sluchays['oslo'] != None:
                        for o, oslo in enumerate(sluchays['oslo']):
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).alignment = Alignment(horizontal="center", vertical="center")
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).number_format = numbers.FORMAT_TEXT
                            if data == 'oslo_tnvr':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oslo.tnvr.kod if oslo.tnvr != None else ''
                            elif data == 'oslo_date':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oslo.dato
                            elif data == 'oslo_kod_osl':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oslo.osl.kod if oslo.osl != None else ''
                            elif data == 'oslo_xosl':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oslo.xosl.naim if oslo.xosl != None else ''
                            elif data == 'oslo_posl':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oslo.posl.naim if oslo.posl != None else ''
                            elif data == 'oslo_aosl':
                                    self.sheet.cell(column=1 + l, row=i + 1 + o+row).value = oslo.aosl.naim if oslo.aosl != None else ''

                    if data == 'trs':
                        self.sheet.cell(column=1 + l,row=i + 1 + row).value = sluchays['sluchay'].trs.naim if sluchays['sluchay'].trs != None else ''

                    if sluchays['manpy'] != None:
                        for o, manpy in enumerate(sluchays['manpy']):
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).alignment = Alignment(horizontal="center", vertical="center")
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).number_format = numbers.FORMAT_TEXT
                            if data == 'man_date':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = manpy.datm
                            elif data == 'man_tnvr':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = manpy.tnvr.kod if manpy.tnvr != None else ''
                            elif data == 'man_kodmn':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = manpy.kodmn.kod if manpy.kodmn != None else ''
                            elif data == 'man_kol':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = manpy.kol
                            elif data == 'man_pl':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = manpy.pl_display() if manpy.pl_display() != None else ''


                    if sluchays['vb_s']!= None:
                        for o,sluch in enumerate(sluchays['vb_s']):
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).alignment = Alignment(horizontal="center", vertical="center")
                            self.sheet.cell(column=1 + l, row=i + 1 + o + row).number_format = numbers.FORMAT_TEXT
                            if data == 'potd':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = sluch.potd.naim if sluch.potd != None else ''
                            elif data == 'dat_pe':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = sluch.dat_pe
                            elif data == 'kod_y':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = sluch.kod_y.naim if sluch.kod_y != None else ''
                            elif data == 'pr_per':
                                self.sheet.cell(column=1 + l,row=i + 1 + o + row).value = sluch.pr_per.naim if sluch.pr_per != None else None

                    if data == 'wskr_date':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].wskr_date
                    elif data == 'tm_let':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].tm_let
                    elif data == 'pri':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].pri.naim if sluchays['sluchay'].pri != None else ''
                    elif data == 'ds_let_kod':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].ds_let.kod if sluchays['sluchay'].ds_let != None else ''
                    elif data == 'wskr':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value =  sluchays['sluchay'].wskr_display() if sluchays['sluchay'].wskr_display() != None else ''
                    elif data == 'dspat_kod':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].dspat.kod if sluchays['sluchay'].dspat != None else ''
                    elif data == 'rasxp':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].rasxp_display() if sluchays['sluchay'].rasxp_display() != None else ''
                    elif data == 'otd_y':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].otd_y.naim if sluchays['sluchay'].otd_y != None else ''

                    if sluchays['le_trv'] != None:
                        # print(sluchays['le_trv'].id)
                        if data == 'dskz_kod':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].dskz.kod if sluchays['sluchay'].dskz != None else ''
                        elif data == 'details_kod':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['le_trv'].details.kod if sluchays['le_trv'].details != None else ''
                        elif data == 't_trv':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['le_trv'].t_trv.naim if sluchays['le_trv'].t_trv != None else ''
                        elif data == 'trav_ns':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['le_trv'].trav_ns.naim if sluchays['le_trv'].trav_ns != None else ''

                    if data == 'vds':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vds'].vds.naim if sluchays['vds'].vds != None else ''
                    elif data == 'sctp':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vds'].sctp
                    elif data == 'nctp':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vds'].nctp
                    elif data == 'ctkom':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vds'].ctkom.naim if sluchays['vds'].ctkom != None else ''
                    elif data == 't_pol':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vds'].t_pol.tip_name if sluchays['vds'].t_pol != None else ''
                    elif data == 'udl':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].udl.docname if sluchays['patient'].udl != None else ''
                    elif data == 's_pasp':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].s_pasp
                    elif data == 'n_pasp':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].n_pasp
                    elif data == 'docdate':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].docdate
                    elif data == 'docorg':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].docorg
                    elif data == 'm_roj':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].m_roj
                    elif data == 'ss':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient'].ss

                    if sluchays['vb_a'] != None:
                        if data == 'vb_a_datv':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vb_a'].datv
                        elif data == 'srber':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vb_a'].srber
                        elif data == 'n_ber':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vb_a'].n_ber
                        elif data == 'pria':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vb_a'].pria.naim if sluchays['vb_a'].pria != None else ''
                        elif data == 'm_prer':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['vb_a'].m_prer.naim if sluchays['vb_a'].m_prer != None else ''

                    if sluchays['disability'] != None:
                        if data == 'dat_l1':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['disability'].dat_l1.strftime('%d-%m-%Y') if sluchays['disability'].dat_l1 != None else ''
                        elif data == 'dat_l2':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['disability'].dat_l2.strftime('%d-%m-%Y') if sluchays['disability'].dat_l2 != None else ''
                        elif data == 'ot_ln':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = 'Да' if sluchays['disability'].ot_ln else 'Нет'
                        elif data == 'vs_bol':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['disability'].vs_bol
                        elif data == 'dis_sex_bol':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['disability'].sex_bol.polname if sluchays['disability'].sex_bol != None else ''

                    if sluchays['patient_p'] != None:
                        if data == 'fam_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].fam_p
                        elif data == 'im_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].im_p
                        elif data == 'ot_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].ot_p
                        elif data == 'pol_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].pol.polname if sluchays['patient_p'].pol != None else ''
                        elif data == 'mp_roj':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].m_roj
                        elif data == 'udl_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].udl_p.docname if sluchays['patient_p'].udl_p != None else ''
                        elif data == 'sp_pasp':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].sp_pasp
                        elif data == 'np_pasp':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].np_pasp
                        elif data == 'skom_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].skom_p.naim if sluchays['patient_p'].skom_p != None else ''
                        elif data == 'stat_p':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].stat_p.tip_name if sluchays['patient_p'].stat_p != None else ''
                        elif data == 's_pol':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].s_pol
                        elif data == 'n_pol':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['patient_p'].n_pol

                    if sluchays['onk_sl'] != None:
                        if data == 'ds1_t':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_sl'].ds1_t.reas_name if sluchays['onk_sl'].ds1_t != None else ''
                        elif data == 'stad':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_sl'].stad.ds_st if sluchays['onk_sl'].stad != None else ''
                        elif data == 'onk_t':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_sl'].onk_t.kod_t if sluchays['onk_sl'].onk_t != None else ''
                        elif data == 'onk_n':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_sl'].onk_n.kod_n if sluchays['onk_sl'].onk_n != None else ''
                        elif data == 'onk_m':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_sl'].onk_m.kod_m if sluchays['onk_sl'].onk_m != None else ''
                        elif data == 'mtstz':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_sl'].mtstz_display() if sluchays['onk_sl'].mtstz != None else ''

                    if data == 'c_zab':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].c_zab.n_cz if sluchays['sluchay'].c_zab != None else ''

                    if sluchays['b_diag'] != None:
                        if data == 'diag_date':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_diag'].diag_date.strftime('%d.%m.%Y') if sluchays['b_diag'].diag_date != None else ''
                        elif data == 'diag_tip':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_diag'].diag_tip_display() if sluchays['b_diag'].diag_tip != None else ''
                        elif data == 'diag_code':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_diag'].diag_code
                        elif data == 'diag_rslt':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_diag'].diag_rslt
                        elif data == 'rec_rslt':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_diag'].rec_rslt
                    if sluchays['cons'] != None:
                        if data == 'dt_cons':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['cons'].dt_cons.strftime('%d.%m.%Y') if sluchays['cons'].dt_cons != None else ''
                        elif data == 'pr_cons':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['cons'].pr_cons.cons_name if sluchays['cons'].pr_cons != None else ''

                    if sluchays['onk_usl'] != None:
                        if data == 'usl_tip':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_usl'].usl_tip.tlech_name if sluchays['onk_usl'].usl_tip != None else ''
                        elif data == 'hir_tip':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['onk_usl'].hir_tip.thir_name if  sluchays['onk_usl'].hir_tip != None else ''

                    if sluchays['b_prot'] != None:
                        if data == 'd_prot':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_prot'].d_prot.strftime('%d.%m.%Y') if sluchays['b_prot'].d_prot != None else ''
                        elif data == 'prot':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['b_prot'].prot.prot_name if sluchays['b_prot'].prot != None else ''

                    if sluchays['napr'] != None:
                        if data == 'naprdate':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['napr'].naprdate.strftime('%d.%m.%Y') if sluchays['napr'].naprdate != None else ''
                        elif data == 'napr_v':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['napr'].napr_v.n_vn if sluchays['napr'].napr_v != None else ''
                        elif data == 'napr_mo':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['napr'].napr_mo.naim if sluchays['napr'].napr_mo != None else ''
                        elif data == 'napr_issl':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['napr'].napr_issl.n_met if sluchays['napr'].napr_issl != None else ''
                        elif data == 'napr_usl':
                            self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['napr'].napr_usl.kod if sluchays['napr'].napr_usl != None else ''

                    if data == 'pmg':
                        self.sheet.cell(column=1 + l, row=i + 1 + row).value = sluchays['sluchay'].pmg.naim if sluchays['sluchay'].pmg != None else ''
                end_row += max(max_row) if len(max_row) != 0 else 0

            self._path()
            # self.wb.save(self._dir + f'g_oth_all_{self._user.user.id}.xlsx')
            self.wb.save(self._dir+self._filename)
            # print(self._dir+self._filename)
            # if os.path.exists(self._dir+self._filename):
            #     with open(self._dir+self._filename,'rb') as fh:
            #         response = HttpResponse(fh.read(),content_type="application/adminupload")
            #         response['Content-Disposition'] = 'inline;filename='+os.path.basename(self._dir+self._filename)
            async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                          {'type': 'download', 'text': self._dir+self._filename})
            # async_to_sync(get_channel_layer().group_send)(self._user_group_name,
            #                                               {'type': 'report_data', 'text': 'Отчет сфромирован'})

    def header_d(self):
        self.header = dict()
        self.header['nib'] = 'История'
        self.header['fam'] = 'Фамилия'
        self.header['im'] = 'Имя'
        self.header['ot'] = 'Отчество'
        self.header['pol'] = 'Пол'
        self.header['datp'] = 'Дата поступления'
        self.header['tm_otd'] = 'Время'
        self.header['datv'] = 'Дата выписки'
        self.header['datr'] = 'Дата рождения'
        self.header['otd'] = 'Отделение'
        self.header['m_roj_in'] = 'Адрес рождения'
        self.header['adr_in'] = 'Адрес регистрации'
        self.header['rab'] = 'Место работы'
        self.header['prof'] = 'Профессия'
        self.header['r_n'] = 'Социальный статус'
        self.header['in_t'] = 'Категория льготности'
        self.header['lpy'] = 'Кем направлен'
        self.header['npr_num'] = '№ Направления'
        self.header['npr_date'] = 'Дата направления'
        self.header['alg'] = 'Подозрениена опьянение'
        self.header['goc'] = 'Госпитализация'
        self.header['prpg'] = 'Обращения'
        self.header['vrez'] = 'Давность заболевания'
        self.header['p_per'] = 'Признак поступления'
        self.header['dsny'] = 'Ds направив.\nучреждения'
        self.header['ds_0'] = 'Ds при\nпоступлении'
        self.header['dsk'] = 'Ds\nКлинический'
        self.header['dskz'] = 'Ds\nКлин.заключ'
        self.header['ds_osl'] = 'Ds\nосложнения'
        self.header['dsc'] = 'Ds\nсопутствующий'
        self.header['dson'] = 'Ds\nонкологий'
        self.header['dat_otd'] = 'Поступления\nиз приемного отд'
        self.header['tm_otd_d'] = 'Время поступления\nиз приемного отд'
        self.header['icx'] = 'Исход\nлечения'
        self.header['rslt'] = 'Результат\nлечения'
        self.header['koy_N'] = 'Всего койко-дни'
        self.header['koy_aro'] = 'В Аро'
        self.header['koy_otd'] = 'В профильном\nотделении'
        self.header['koy_prof_k'] = 'Отделение'
        self.header['koy_kod'] = 'Код врача'
        self.header['oper_date'] = 'Дата операции'
        self.header['oper_tm'] = 'Время операции'
        self.header['oper_py'] = 'ПО/СТАЦ'
        self.header['oper_kod_op'] = 'Код операции'
        self.header['oper_goc'] = 'ПЛ/ЭК'
        self.header['oper_kodx'] = 'Код хирурга'
        self.header['oper_pop'] = 'Основ.опер'
        self.header['oper_pr_osob'] = 'Особ-ти операции'
        self.header['oper_k_mm'] = 'К-во.биом'
        self.header['oper_kodxa'] = '1-Ассистент'
        self.header['oper_kodxa1'] = '2-Ассистент'
        self.header['oper_obz'] = 'Метод.обез-я'
        self.header['oper_kodan'] = 'Анестезиолог'
        self.header['ksg_osn'] = 'КСГ осн.заболевания'
        self.header['oopkk'] = 'Классификатор критерии'
        self.header['ksg_sop'] = 'КСГ сопут.зап'
        self.header['iddoc'] = 'Врач'
        self.header['oslo_tnvr'] = 'Таб.N'
        self.header['oslo_date'] = 'Дата осложнения'
        self.header['oslo_kod_osl'] = 'Код осложнения'
        self.header['oslo_xosl'] = 'Характер осл.'
        self.header['oslo_posl'] = 'Причина'
        self.header['oslo_aosl'] = 'Экспертиза'
        self.header['trs'] = 'Трудоспособность'
        self.header['man_date'] = 'Дата манипуляции'
        self.header['man_tnvr'] = 'Таб.N '
        self.header['man_kodmn'] = 'Код Манипуляции'
        self.header['man_kol'] = 'К-во'
        self.header['man_pl'] = 'Плат.Услуга'
        self.header['potd'] = 'Перевод из Отд'
        self.header['dat_pe'] = 'Дата перевода'
        self.header['kod_y'] = 'Перевод в Др.ЛПУ'
        self.header['pr_per'] = 'Причина перевода'
        self.header['wskr_date'] = 'Дата смерти'
        self.header['tm_let'] = 'Время смерти'
        self.header['pri'] = 'Код причины\nлетального исхода'
        self.header['ds_let_kod'] = 'Ds причины\nлетального исхода'
        self.header['wskr'] = 'Вскрытие'
        self.header['dspat_kod'] = 'Паталогоанатомический'
        self.header['rasxp'] = 'Расхождение'
        self.header['otd_y'] = 'Умер в'
        self.header['dskz_kod'] = 'Характер травмы'
        self.header['details_kod'] = 'Внешние причины\nтравмы'
        self.header['t_trv'] = 'Тип травмы'
        self.header['trav_ns'] = 'Третьих лиц'
        self.header['vds'] = 'Источник оплаты'
        self.header['sctp'] = 'Сер.полиса'
        self.header['nctp'] = 'N полиса'
        self.header['ctkom'] = 'СМО'
        self.header['t_pol'] = 'Тип полиса'
        self.header['udl'] = 'Тип документа'
        self.header['s_pasp'] = 'Сер.ДУЛ'
        self.header['n_pasp'] = 'N ДУЛ'
        self.header['docdate'] = 'Дата выдачи'
        self.header['docorg'] = 'Кем выдан'
        self.header['m_roj'] = 'Место рождения'
        self.header['ss'] = 'Снилс'
        self.header['vb_a_datv'] = 'Дата'
        self.header['srber'] = 'Срок беременности\n(недель)'
        self.header['n_ber'] = 'Настоящая беременность\nпо счету'
        self.header['pria'] = 'Причины прерывания\nбеременности'
        self.header['m_prer'] = 'Метод прерывания\nбеременности'
        self.header['dat_l1'] = 'Открыт с'
        self.header['dat_l2'] = 'Закрыт по'
        self.header['ot_ln'] = 'Закрыт'
        self.header['vs_bol'] = 'Полных лет'
        self.header['dis_sex_bol'] = 'Пол'
        self.header['fam_p'] = 'Фимилия'
        self.header['im_p'] = 'Имя'
        self.header['ot_p'] = 'Отчество'
        self.header['pol_p'] = 'Пол'
        self.header['mp_roj'] = 'Место рождения'
        self.header['udl_p'] = 'Тип документа'
        self.header['sp_pasp'] = 'Серия'
        self.header['np_pasp'] = 'Номер'
        self.header['skom_p'] = 'СМО'
        self.header['stat_p'] = 'Тип полиса'
        self.header['s_pol'] = 'Сер.полиса'
        self.header['n_pol'] = 'N.полиса'
        self.header['ds1_t'] = 'Повод обращения'
        self.header['stad'] = 'Стадия заболевания'
        self.header['onk_t'] = 'Стадия по T'
        self.header['onk_n'] = 'Стадия по N'
        self.header['onk_m'] = 'Стадия по M'
        self.header['mtstz'] = 'Наличие отдельных\nметастазов'
        self.header['c_zab'] = 'Характер заболевания'
        self.header['diag_date'] = 'Дата'
        self.header['diag_tip'] = 'Тип диагностического\nпоказателя'
        self.header['diag_code'] = 'Код диагностического\nпоказателя'
        self.header['diag_rslt'] = 'Код результата\nдиагностики'
        self.header['rec_rslt'] = 'Признак получения\nрезультата диагностики'
        self.header['dt_cons'] = 'Дата'
        self.header['pr_cons'] = 'Цель консилиума'
        self.header['usl_tip'] = 'Тип услуги'
        self.header['hir_tip'] = 'Тип хирургического\nлечения'
        self.header['d_prot'] = 'Дата'
        self.header['prot'] = 'Противопоказания \nи отказы'
        self.header['naprdate'] = 'Дата направления'
        self.header['napr_v'] = 'Вид направления'
        self.header['napr_mo'] = 'Мо,куда рекомендовано\nобращаться'
        self.header['napr_issl'] = 'Метод диагностич.\nисследования'
        self.header['napr_usl'] = 'Мед.услуга(код)\nрекомендованная'
        self.header['pmg'] = 'Мо прикрепления'

        self.header['nib_width'] = 12
        self.header['fam_width'] = 18
        self.header['im_width'] = 18
        self.header['ot_width'] = 18
        self.header['pol_width'] = 12
        self.header['datp_width'] = 18
        self.header['tm_otd_width'] = 12
        self.header['datv_width'] = 18
        self.header['datr_width'] = 18
        self.header['otd_width'] = 25
        self.header['m_roj_in_width'] = 80
        self.header['adr_in_width'] = 80
        self.header['rab_width'] = 35
        self.header['prof_width'] = 35
        self.header['r_n_width'] = 35
        self.header['in_t_width'] = 35
        self.header['lpy_width'] = 40
        self.header['npr_num_width'] = 25
        self.header['npr_date_width'] =18
        self.header['alg_width'] = 25
        self.header['goc_width'] = 25
        self.header['prpg_width'] = 25
        self.header['vrez_width'] = 25
        self.header['p_per_width'] = 25
        self.header['dsny_width'] = 12
        self.header['ds_0_width'] = 12
        self.header['dsk_width'] = 12
        self.header['dskz_width'] = 12
        self.header['ds_osl_width'] = 12
        self.header['dsc_width'] = 12
        self.header['dson_width'] = 12
        self.header['dat_otd_width'] = 18
        self.header['tm_otd_d_width'] = 18
        self.header['icx_width'] = 20
        self.header['rslt_width'] = 20
        self.header['koy_N_width'] = 16
        self.header['koy_aro_width'] = 16
        self.header['koy_otd_width'] = 16
        self.header['koy_prof_k_width'] = 25
        self.header['koy_kod_width'] = 16
        self.header['oper_date_width'] = 18
        self.header['oper_tm_width'] = 16
        self.header['oper_py_width'] = 28
        self.header['oper_kod_op_width'] = 18
        self.header['oper_goc_width'] = 16
        self.header['oper_kodx_width'] = 18
        self.header['oper_pop_width'] = 16
        self.header['oper_pr_osob_width'] = 18
        self.header['oper_k_mm_width'] = 16
        self.header['oper_kodxa_width'] = 16
        self.header['oper_kodxa1_width'] = 16
        self.header['oper_obz_width'] = 16
        self.header['oper_kodan_width'] = 16
        self.header['ksg_osn_width'] = 22
        self.header['oopkk_width'] = 22
        self.header['ksg_sop_width'] = 16
        self.header['iddoc_width'] = 16
        self.header['oslo_tnvr_width'] = 18
        self.header['oslo_date_width'] = 18
        self.header['oslo_kod_osl_width'] = 18
        self.header['oslo_xosl_width'] = 35
        self.header['oslo_posl_width'] = 22
        self.header['oslo_aosl_width'] = 35
        self.header['trs_width'] = 30
        self.header['man_date_width'] = 18
        self.header['man_tnvr_width'] = 18
        self.header['man_kodmn_width'] = 18
        self.header['man_kol_width'] = 18
        self.header['man_pl_width'] = 18
        self.header['potd_width'] = 18
        self.header['dat_pe_width'] = 18
        self.header['kod_y_width'] = 40
        self.header['pr_per_width'] = 40
        self.header['wskr_date_width'] = 18
        self.header['tm_let_width'] = 18
        self.header['pri_width'] = 18
        self.header['ds_let_kod_width'] = 18
        self.header['wskr_width'] = 18
        self.header['dspat_kod_width'] = 18
        self.header['rasxp_width'] = 18
        self.header['otd_y_width'] = 18
        self.header['dskz_kod_width'] = 18
        self.header['details_kod_width'] = 18
        self.header['t_trv_width'] = 50
        self.header['trav_ns_width'] = 50
        self.header['vds_width'] = 35
        self.header['sctp_width'] = 20
        self.header['nctp_width'] = 20
        self.header['ctkom_width'] = 40
        self.header['t_pol_width'] = 40
        self.header['udl_width'] = 40
        self.header['s_pasp_width'] = 18
        self.header['n_pasp_width'] = 18
        self.header['docdate_width'] = 18
        self.header['docorg_width'] = 18
        self.header['m_roj_width'] = 80
        self.header['ss_width'] = 18
        self.header['vb_a_datv_width'] = 18
        self.header['srber_width'] = 22
        self.header['n_ber_width'] = 22
        self.header['pria_width'] = 22
        self.header['m_prer_width'] = 22
        self.header['dat_l1_width'] = 18
        self.header['dat_l2_width'] = 18
        self.header['ot_ln_width'] = 18
        self.header['vs_bol_width'] = 18
        self.header['dis_sex_bol_width'] = 18
        self.header['fam_p_width'] = 22
        self.header['im_p_width'] = 22
        self.header['ot_p_width'] = 22
        self.header['pol_p_width'] = 18
        self.header['mp_roj_width'] = 80
        self.header['udl_p_width'] = 40
        self.header['sp_pasp_width'] = 18
        self.header['np_pasp_width'] = 18
        self.header['skom_p_width'] = 45
        self.header['stat_p_width'] = 45
        self.header['s_pol_width'] = 30
        self.header['n_pol_width'] = 22
        self.header['ds1_t_width'] = 24
        self.header['stad_width'] = 18
        self.header['onk_t_width'] = 18
        self.header['onk_n_width'] = 18
        self.header['onk_m_width'] = 18
        self.header['mtstz_width'] = 20
        self.header['c_zab_width'] = 45
        self.header['diag_date_width'] = 18
        self.header['diag_tip_width'] = 30
        self.header['diag_code_width'] = 40
        self.header['diag_rslt_width'] = 40
        self.header['rec_rslt_width'] = 24
        self.header['dt_cons_width'] = 18
        self.header['pr_cons_width'] = 36
        self.header['usl_tip_width'] = 50
        self.header['hir_tip_width'] = 60
        self.header['d_prot_width'] = 18
        self.header['prot_width'] = 50
        self.header['naprdate_width'] = 18
        self.header['napr_v_width'] = 50
        self.header['napr_mo_width'] = 60
        self.header['napr_issl_width'] = 80
        self.header['napr_usl_width'] = 24
        self.header['pmg_width'] = 80
