from services.hospital.reports import Reports
from services.hospital.patient import Patients
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from openpyxl.styles import numbers
from openpyxl import Workbook
from  openpyxl.styles import Font,Alignment,Border,Side
from openpyxl import load_workbook
import json
import copy
import os
import shutil
from django.conf import  settings
from okb2.models import MyUser
from celery.contrib import rdb
from hospital.models import *
import datetime

class Create(Reports):

    def __init__(self,user,request):
        super().__init__(user,request)
        
        self._user_group_name = 'hospital_reports_%s' % user
        # self.list_data = json.loads(self._request['list_data']) if self._request.get('list_data') != None else []
        self._path_shoblons = 'shoblons/hospital/oth'
        self._temp_dir = 'temp'
        self._border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
    def create(self):
        self._sluchays = self.get_sluchays()
        self.old_sluchays = copy.deepcopy(self._sluchays)

        #Допилить с фильтрами
        if self._request.get('filters',None) != None:
            filters = json.loads(self._request.get('filters'))
            if filter:
                self._sluchays = self.filter(filters, self._sluchays)




        type_fun = self._request.get('type_report')
        self._filename = self._request.get('filename', None)
        self._filename = f'{self._filename}.xlsx' if self._filename != None and len(self._filename) > 0 else f'{type_fun}_{self._user.user.id}.xlsx'
        if type_fun == 'a_oth_1':
            self._a_oth_1()
        elif type_fun == 'a_oth_2':
            self._a_oth_2()
        elif type_fun == 'a_oth_3':
            self._a_oth_3()
        elif type_fun == 'a_oth_4':
            self._a_oth_4(self._sluchays)
        elif type_fun == 'a_oth_5':
            self._a_oth_5(self._sluchays,self.old_sluchays)
        elif type_fun == 'a_oth_6':
            self._a_oth_6(self._sluchays)
        elif type_fun == 'a_oth_7':
            self._a_oth_7()
        elif type_fun == 'a_oth_8':
            self._a_oth_8(self._sluchays)
        elif type_fun == 'a_oth_9':
            self._a_oth_9()
        elif type_fun == 'a_oth_10':
            self._a_oth_10()
        elif type_fun == 'a_oth_11':
            self._a_oth_11(self._sluchays)
        elif type_fun == 'a_oth_12':
            self._a_oth_12(self._sluchays)
        elif type_fun == 'a_oth_13':
            self._a_oth_13(self._sluchays)
        elif type_fun == 'a_oth_14':
            self._a_oth_14()
        elif type_fun == 'a_oth_15':
            self._a_oth_15()
        elif type_fun == 'a_oth_16':
            self._a_oth_16()
        elif type_fun == 'a_oth_17':
            self._a_oth_17()
        elif type_fun == 'a_oth_18':
            self._a_oth_18()
        elif type_fun == 'a_oth_19':
            self._a_oth_19()
        elif type_fun == 'a_oth_20':
            self._a_oth_20()
        elif type_fun == 'a_oth_21':
            self._a_oth_21()
        elif type_fun == 'a_oth_22':
            self._a_oth_22()
        elif type_fun == 'a_oth_23':
            self._a_oth_23()
        elif type_fun == 'a_oth_24':
            self._a_oth_24()
        elif type_fun == 'a_oth_25':
            self._a_oth_25()
        elif type_fun == 'a_oth_26':
            self._a_oth_26()
        elif type_fun == 'a_oth_27':
            self._a_oth_27()
        elif type_fun == 'a_oth_28':
            self._a_oth_28()
        elif type_fun == 'a_oth_29':
            self._a_oth_29()
        elif type_fun == 'a_oth_30':
            self._a_oth_30()
        elif type_fun == 'a_oth_31':
            self._a_oth_31()
        elif type_fun == 'a_oth_32':
            self._a_oth_32()
        elif type_fun == 'a_oth_33':
            self._a_oth_33()
        elif type_fun == 'a_oth_34':
            self._a_oth_34()

  
    def _a_oth_1(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT,self._path_shoblons,'a_oth_1.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT,self._temp_dir,str(self._user.id),file_new])
            shutil.copy2(file_shoblon,file_new)
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self._date_1.strftime("%d.%m.%Y")} по {self._date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self._user.statistics_type.name).capitalize()
            filetrs = ['otd','manpy','goc_ek','goc_pl']
            temp = self.filetr_sluchays(filetrs)
            res_set = [[r, 0, 0, 0] for r in list(set(otd[0] for otd in temp))]
            row = 9
            for t in temp:
                for r in res_set:
                    if r[0] == t[0]:
                        r[1] += t[1]
                        r[2] += t[2]
                        r[3] += t[3]
            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self._border
            wb.save(file_new)
        #     # if self.user.statistics_type.id == 2:
        #     #     shutil.copy2(file_new,'/'.join([settings.MEDIA_MNT,'stat','STW','Отчёты']))
        #     # elif self.user.statistics_type.id == 1:
        #     #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT,self._path_shoblons,'a_oth_2.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon,file_new)
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self._date_1.strftime("%d.%m.%Y")} по {self._date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self._user.statistics_type.name).capitalize()
            filetrs = ['otd','adr_not_tum_obl_rn','adr_tu men','adr_tum_obl_rn','adr_tumen_rn','adr_hm','adr_im','not_obl_rn_tum','oksm_not_643']
            temp = self.filetr_sluchays(filetrs)
            # for slych in self._sluchays:
            #     temp.append(
            #         [
            #             slych['sluchay'].otd.naim if slych['sluchay'].otd != None else '',
            #             1 if slych['patient'].adr != '' and ((('Тюменская обл' not in slych['patient'].adr) and (
            #                         'обл. Тюменская' not in slych['patient'].adr) and ('ОБЛ ТЮМЕНСКАЯ' not in slych['patient'].adr)) \
            #                                                  and (('Тюменский р-н' not in slych['patient'].adr) and (
            #                                 'р-н. Тюменский' not in slych['patient'].adr))) else 0,
            #             1 if slych['patient'].adr != '' and 'Тюмень' in slych['patient'].adr else 0,
            #             1 if slych['patient'].adr != '' and ((('Тюменская обл' in slych['patient'].adr) or (
            #                         'обл. Тюменская' in slych['patient'].adr) or ('ОБЛ ТЮМЕНСКАЯ' in slych['patient'].adr)) \
            #                                                  and (('Тюменский р-н' not in slych['patient'].adr) or (
            #                                 'р-н. Тюменский' not in slych['patient'].adr))) else 0,
            #             1 if slych['patient'].adr != '' and 'Тюменский р-н' in slych[
            #                 'patient'].adr or 'р-н. Тюменский' in slych['patient'].adr else 0,
            #             1 if slych['patient'].adr != '' and 'Ханты-Мансийский' in slych['patient'].adr else 0,
            #             1 if slych['patient'].adr != '' and 'Ямало-Ненецкий' in slych['patient'].adr else 0,
            #             1 if slych['patient'].adr != '' and ((('Тюменская обл' not in slych['patient'].adr) and (
            #                         'обл. Тюменская' not in slych['patient'].adr) and ('ОБЛ ТЮМЕНСКАЯ' not in slych['patient'].adr)) \
            #                                                  and (('Тюменский р-н' not in slych['patient'].adr) and ('р-н. Тюменский' not in slych['patient'].adr))) else 0,
            #             1 if slych['patient'].c_oksm != None and slych['patient'].c_oksm.kod != 643 else 0
            #         ]
            #     )

            res_set = [[r, 0, 0, 0, 0, 0, 0, 0, 0, 0] for r in list(set(otd[0] for otd in temp))]

            row = 9
            for t in temp:
                for r in res_set:
                    if r[0] == t[0]:
                        r[1] += sum(t[1:])
                        r[2] += t[2]
                        r[3] += t[3]
                        r[4] += t[4]
                        r[5] += t[5]
                        r[6] += t[6]
                        r[7] += t[7]
                        r[8] += t[8]

            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).border = self._border
                    sheet.cell(row=i + row,column=1 + l).alignment = Alignment(horizontal="center",vertical="center")

            wb.save(file_new)
            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_3(self):
        pass
    def _a_oth_4(self,sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_4.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self._user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_5(self,sluchays,old_sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_5.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            old_sluchays = old_sluchays
            temp_filter = []
            temp_old = []
           

            for sluchay in sluchays:
                pop = Oper.objects.filter(pop=True,sluchay__id=sluchay['sluchay'].id)
                temp_filter.append([
                    sluchay['sluchay'].otd.naim if sluchay['sluchay'].otd != None else '',
                    1,
                    1 if sluchay['sluchay'].goc != None and sluchay['sluchay'].goc.tip_name == 'Экстренная' else 0,
                    sluchay['le_vr'].kd if sluchay['le_vr'] !=  None  and sluchay['le_vr'].kd != None else 0,
                    #Средние значение нужно доделать когда будут записи
                    0,
                    1 if len(pop)>0 else 0,
                    1 if sluchay['sluchay'].rslt != None and ((sluchay['sluchay'].rslt.tip_name == 'Умер 105') or
                                                    (sluchay['sluchay'].rslt.tip_name == 'Умер в приёмном покое 106')) else 0,
                    (sluchay['sluchay'].datv - pop[0].dato).days if ((len(pop)>0) and (pop[0].dato != None)) else 0,
                    ])
            
            for sluchay in old_sluchays:
                pop = Oper.objects.filter(pop=True,sluchay__id=sluchay['sluchay'].id)
                temp_old.append([
                    sluchay['sluchay'].otd.naim if sluchay['sluchay'].otd != None else '',
                    1,
                    1 if sluchay['sluchay'].goc != None and sluchay['sluchay'].goc.tip_name == 'Экстренная' else 0,
                    sluchay['le_vr'].kd if sluchay['le_vr'] !=  None  and sluchay['le_vr'].kd != None else 0,
                    #Средние значение нужно доделать когда будут записи
                    0,
                    1 if len(pop)>0 else 0,
                    1 if sluchay['sluchay'].rslt != None and ((sluchay['sluchay'].rslt.tip_name == 'Умер 105') or
                                                    (sluchay['sluchay'].rslt.tip_name == 'Умер в приёмном покое 106')) else 0,
                    (sluchay['sluchay'].datv - pop[0].dato).days if ((len(pop)>0) and (pop[0].dato != None)) else 0,
                    ])
                    
            res_set_filter = [[r, 0, 0, 0, 0, 0, 0, 0] for r in list(set(otd[0] for otd in temp_filter))]
           
            for t in temp_filter:
                for r in res_set_filter:
                    if r[0] == t[0]:
                        r[1] += t[1]
                        r[2] += t[2]
                        r[3] += t[3]
                        r[4] += t[4]
                        r[5] += t[5]
                        r[6] += t[6]
                        r[7] += t[7]

            res_set_old = [[r, 0, 0, 0, 0, 0, 0, 0] for r in list(set(otd[0] for otd in temp_old))]
           
            for t in temp_filter:
                for r in res_set_old:
                    if r[0] == t[0]:
                        r[1] += t[1]
                        r[2] += t[2]
                        r[3] += t[3]
                        r[4] += t[4]
                        r[5] += t[5]
                        r[6] += t[6]
                        r[7] += t[7]
            
            if len(res_set_filter) == len(res_set_old):
                res_new = [[
                        res_set_old[r][0],
                        res_set_old[r][1],
                        res_set_filter[r][1],
                        res_set_old[r][2],
                        res_set_filter[r][2],
                        res_set_old[r][3],
                        res_set_filter[r][3],
                        res_set_old[r][4],
                        res_set_filter[r][4],
                        res_set_old[r][5],
                        res_set_filter[r][5],
                        res_set_old[r][6],
                        res_set_filter[r][6],
                        res_set_old[r][7],
                        res_set_filter[r][7],
                        ] for r in range(len(res_set_filter))]                     
     
            # rdb.set_trace()
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self._user.statistics_type.name).capitalize()
            row = 9
            for i, res in enumerate(res_new):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).border = self.border
                    sheet.cell(row=i + row,column=1 + l).alignment = Alignment(horizontal="center",vertical="center")
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_6(self,sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_6.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            #взрос >= 18
            #Дети <10
            #Подростки 10-16
            # rdb.set_trace()
            #Нужны данные 


            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self._user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_7(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_8(self,sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_8.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            temp = []
            for slych in self._sluchays:
                goc = slych['sluchay'].goc
                datp = slych['sluchay'].datp
                datv = slych['sluchay'].datv
                date_raz = datv - datp
                if goc != None and goc.tip_name == 'Экстренная':  
                    temp.append(
                        [
                            slych['sluchay'].otd.naim if slych['sluchay'].otd != None else '',
                            1 if goc != None and goc.tip_name == 'Экстренная' else 0,
                            1 if ((goc != None and goc.tip_name == 'Экстренная') and (1<=date_raz.days<=3))  else 0,
                        ]
                    )
            res_set = [[r, 0, 0] for r in list(set(otd[0] for otd in temp))]
            for t in temp:
                for r in res_set:
                    if r[0] == t[0]:
                        r[1] += t[1]
                        r[2] += t[2]

            # rdb.set_trace()
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self._user.statistics_type.name).capitalize()
            row = 8
            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self.border
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_9(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_9.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            #Нужно понять что значит степень тяжести
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_10(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_11(self,sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_11.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            #Нужны данные 
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_12(self,sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_12.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            temp = []
            for sluchay in sluchays:
                datp = sluchay['sluchay'].datp
                datv = sluchay['sluchay'].datv
                date_raz = datv - datp
                rslt = sluchay['sluchay'].rslt
                if rslt != None and ((rslt.tip_name == 'Умер 105') or(rslt.tip_name == 'Умер в приёмном покое 106')):
                    temp.append(
                        [
                            sluchay['sluchay'].otd.naim if sluchay['sluchay'].otd != None else '',
                            1,
                            1 if ((rslt.tip_name == 'Умер 105') or(rslt.tip_name == 'Умер в приёмном покое 106')) else 0,
                            1 if date_raz.days < 1 else 0,
                            1 if date_raz.days < 3 else  0,
                            # Реанимация ???
                            0
                        ]
                    )
            res_set = [[r,0,0,0,0,0] for r in list(set(otd[0] for otd in temp))]
            for t in temp:
                for r in res_set:
                    if r[0] == t[0]:
                        r[1] += t[1]
                        r[2] += t[2]
                        r[3] += t[3]
                        r[4] += t[4]
                        r[5] += t[5]
                        
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            row = 9
            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self.border
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_13(self,sluchays):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_13.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            sluchays = sluchays
            temp_ext = []
            temp_neot = []
            temp_pl = []

            for sluchay in sluchays:
                goc = sluchay['sluchay'].goc
                if goc != None and goc.tip_name == 'Экстренная':
                    temp_ext.append(
                            [
                                sluchay['sluchay'].otd.naim if sluchay['sluchay'].otd != None else '',
                                1,
                                sluchay['le_vr'].aro if sluchay['le_vr'] !=  None  and sluchay['le_vr'].aro != None else 0,

                            ]
                        )
                if goc != None and goc.tip_name == 'Неотложная':
                    temp_neot.append(
                            [
                                sluchay['sluchay'].otd.naim if sluchay['sluchay'].otd != None else '',
                                1
                            ]
                        )
                if goc != None and goc.tip_name == 'Плановая':
                    temp_pl.append(
                            [
                                sluchay['sluchay'].otd.naim if sluchay['sluchay'].otd != None else '',
                                1
                            ]
                        )
                
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_14(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_14.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_15(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_16(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_17(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_18(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_18.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_19(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_19.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            sheet.cell(row=26,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=27, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_20(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_21(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_22(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_23(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_23.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_24(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_24.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_25(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_25.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_26(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_26.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_27(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_28(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_29(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_29.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_30(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_30.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_31(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, 'a_oth_31.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)

            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            wb.save(file_new)

            # if self.user.statistics_type.id == 2:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STW', 'Отчёты']))
            # elif self.user.statistics_type.id == 1:
            #     shutil.copy2(file_new, '/'.join([settings.MEDIA_MNT, 'stat', 'STD', 'Отчёты']))

            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_32(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_33(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})
    def _a_oth_34(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет сфромирован'})

    # def _filetr(self,filters,sluchays_old=None):
    #     rez = []
    #     if len(filters) > 0:
    #         if sluchays_old != None:
    #             sluchays = sluchays_old.copy()
    #         else:
    #             sluchays = self._sluchays.copy()

    #         for slych in sluchays:
    #             temp = []
    #             for filter in filters:
    #                 if filter == 'otd':
    #                     temp.append(slych['sluchay'].otd.naim if slych['sluchay'].otd != None else '')
    #                 elif filter == 'manpy':
    #                     temp.append(len(slych['manpy']) if slych['manpy'] != None else 0)
    #                 elif filter == 'goc_ek':
    #                     temp.append(1 if slych['sluchay'].goc != None and slych['sluchay'].goc.tip_name == 'Экстренная' else 0)
    #                 elif filter == 'goc_pl':
    #                     temp.append(1 if slych['sluchay'].goc != None and slych['sluchay'].goc.tip_name == 'Плановая' else 0)
    #                 elif filter == 'adr_not_tum_obl_rn':
    #                     temp.append(1 if slych['patient'].adr != '' and ((('Тюменская обл' not in slych['patient'].adr) and (
    #                         'обл. Тюменская' not in slych['patient'].adr) and ('ОБЛ ТЮМЕНСКАЯ' not in slych['patient'].adr)) \
    #                         and (('Тюменский р-н' not in slych['patient'].adr) and ('р-н. Тюменский' not in slych['patient'].adr))) else 0)
    #                 elif filter == 'adr_tumen':
    #                     temp.append(1 if slych['patient'].adr != '' and 'Тюмень' in slych['patient'].adr else 0)
    #                 elif filter == 'adr_tum_obl_rn':
    #                     temp.append(1 if slych['patient'].adr != '' and ((('Тюменская обл' in slych['patient'].adr) or (
    #                         'обл. Тюменская' in slych['patient'].adr) or ('ОБЛ ТЮМЕНСКАЯ' in slych['patient'].adr)) \
    #                         and (('Тюменский р-н' not in slych['patient'].adr) or ('р-н. Тюменский' not in slych['patient'].adr))) else 0)
    #                 elif filter == 'adr_tumen_rn':
    #                     temp.append(1 if slych['patient'].adr != '' and 'Тюменский р-н' in slych[
    #                                 'patient'].adr or 'р-н. Тюменский' in slych['patient'].adr else 0)
    #                 elif filter == 'adr_hm':
    #                     temp.append(1 if slych['patient'].adr != '' and 'Ханты-Мансийский' in slych['patient'].adr else 0)
    #                 elif filter == 'adr_im':
    #                     temp.append(1 if slych['patient'].adr != '' and 'Ямало-Ненецкий' in slych['patient'].adr else 0)
    #                 elif filter == 'not_obl_rn_tum':
    #                     temp.append(1 if slych['patient'].adr != '' and ((('Тюменская обл' not in slych['patient'].adr) and (
    #                                 'обл. Тюменская' not in slych['patient'].adr) and ('ОБЛ ТЮМЕНСКАЯ' not in slych['patient'].adr)) \
    #                                 and (('Тюменский р-н' not in slych['patient'].adr) and ('р-н. Тюменский' not in slych['patient'].adr))) else 0)
    #                 elif filter == 'oksm_not_643':
    #                     temp.append(1 if slych['patient'].c_oksm != None and slych['patient'].c_oksm.kod != 643 else 0)


    #             rez.append(temp)
    #     return rez


