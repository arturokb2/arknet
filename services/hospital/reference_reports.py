import copy

from services.hospital.annual_reports import AnnualReportABC, get_rez_apr_1
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from openpyxl import load_workbook,styles
import os
from datetime import datetime
from services.hospital.patient import PatientsData
from collections import OrderedDict
import json
import numpy
from services.hospital.reports import *
from operator import itemgetter, attrgetter, methodcaller
from hospital.models import Oper

border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))


def get_pop_oper(patient):
    pop = patient.sluchay.oper.filter(pop=True) if patient.sluchay.oper else 0
    if pop.count() > 0:
        return pop[0]
    else:
        return None

def get_opers(patient):
    if patient.sluchay.oper.count() > 0:
        opers = patient.sluchay.oper.values('id')
        return [Oper.objects.get(id=o['id']) for o in opers]
    else:
        return None

def get_list_otd(data):
    otds_list = []
    for d in data:
        if d.sluchay.otd.naim not in otds_list:
            otds_list.append(d.sluchay.otd.naim)

    otds_sl = []
    for o in otds_list:
        temp = [[],[]]
        temp[0].append(o)
        for d in data:
            if o == d.sluchay.otd.naim:
                temp[1].append(d)
        otds_sl.append(temp)
    return otds_sl

def get_list_otd_prof(data):
    otds_list = []
    for d in data:
        if d.sluchay.otd.naim not in otds_list:
            otds_list.append(d.sluchay.otd.naim)

    otds_prof_list = []
    for o in otds_list:
        temp = [[],[]]
        temp[0].append(o)
        for d in data:
            if o == d.sluchay.otd.naim:
                if d.sluchay.le_vr and d.sluchay.le_vr.prof_k:
                    if d.sluchay.le_vr.prof_k.k_prname not in temp[1]:
                        temp[1].append(d.sluchay.le_vr.prof_k.k_prname)
        otds_prof_list.append(temp)

    for otds_prof in otds_prof_list:
        for o in range(len(otds_prof[1])):
            otds_prof[1][o] = [otds_prof[1][o],[]]

    for otds_prof in range(len(otds_prof_list)):
        for d in data:
            if otds_prof_list[otds_prof][0][0] == d.sluchay.otd.naim:
                if d.sluchay.le_vr and d.sluchay.le_vr.prof_k:
                    for p in range(len(otds_prof_list[otds_prof][1])):
                        if otds_prof_list[otds_prof][1][p][0] == d.sluchay.le_vr.prof_k.k_prname:
                            otds_prof_list[otds_prof][1][p][1].append(d)

    return otds_prof_list


class PatientsDataFiltrs(PatientsData):
    def __init__(self,date_1,date_2,user,request):
        super().__init__(date_1,date_2,user)
        pats = Reports(user.id,request)
        pats.get_sluchays()

        filters = json.loads(request.get('filters'))
        sl = pats.filter(filters['filter'],pats.patients)
        for s in sl:
            self.get_data(s['sluchay'].id)

def insert_sheet_P1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    count = kwargs['count']
    ym_count = kwargs['ym_count']

    sheet.cell(row=3, column=1).value = str(name).capitalize()
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    row_ = 7
    font = styles.Font(size=16, name='Arial')
    for dat in data:
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:J{sheet.cell(row=row_, column=9).row}")
        sheet.cell(row=row_, column=1).value = dat[0]
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).border = styles.Border(
            bottom=styles.Side(border_style='thin', color='000000'))
        for d in dat[1]:
            row_ += 1
            sheet.row_dimensions[row_].height = 33
            if (row_ % 64 == 0 and row_ <= 64) or (row_ > 64 and row_ % 61 == 3):
                sheet.cell(row=row_, column=1).value = 'История'
                sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=1).font = font
                sheet.cell(row=row_, column=2).value = 'Фамилия.И.О'
                sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=2).font = font
                sheet.cell(row=row_, column=3).value = 'Дата рождения'
                sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=3).font = font
                sheet.cell(row=row_, column=4).value = 'Дата поступ.'
                sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=4).font = font
                sheet.cell(row=row_, column=5).value = 'Дата выбытия'
                sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=5).font = font
                sheet.cell(row=row_, column=6).value = 'Исх леч'
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center",
                                                                            wrap_text=True)
                sheet.cell(row=row_, column=6).font = font
                sheet.cell(row=row_, column=7).value = 'Ds'
                sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=7).font = font
                sheet.merge_cells(
                    f"H{sheet.cell(row=row_, column=8).row}:I{sheet.cell(row=row_, column=9).row}")
                sheet.cell(row=row_, column=8).value = 'Профиль койки'
                sheet.cell(row=row_, column=8).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=8).font = font
                sheet.cell(row=row_, column=10).value = 'Оплат.'
                sheet.cell(row=row_, column=10).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=10).font = font
                row_ += 1
                sheet.row_dimensions[row_].height = 20
                sheet.merge_cells(
                    f"A{sheet.cell(row=row_, column=1).row}:J{sheet.cell(row=row_, column=9).row}")
                sheet.cell(row=row_, column=1).value = dat[0]
                sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center",
                                                                            vertical="center")
                sheet.cell(row=row_, column=1).font = font
                sheet.cell(row=row_, column=1).border = styles.Border(
                    bottom=styles.Side(border_style='thin', color='000000'))
                row_ += 1
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=1).value = d['nib']
            sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center")
            sheet.cell(row=row_, column=1).font = font
            sheet.cell(row=row_, column=2).value = d['fio']
            sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="left", vertical="center")
            sheet.cell(row=row_, column=2).font = font
            sheet.cell(row=row_, column=3).value = d['datr']
            sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=3).font = font
            sheet.cell(row=row_, column=4).value = d['datp']
            sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=4).font = font
            sheet.cell(row=row_, column=5).value = d['datv']
            sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=5).font = font
            sheet.cell(row=row_, column=6).value = d['isx']
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center",
                                                                        wrap_text=True)
            sheet.cell(row=row_, column=6).font = font
            sheet.cell(row=row_, column=7).value = d['dskz']
            sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=7).font = font
            sheet.merge_cells(f"H{sheet.cell(row=row_, column=8).row}:I{sheet.cell(row=row_, column=9).row}")
            sheet.cell(row=row_, column=8).value = d['prof_k']
            sheet.cell(row=row_, column=8).font = font
            sheet.cell(row=row_, column=10).value = d['vds']
            sheet.cell(row=row_, column=10).font = font
    else:
        row_ += 3
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
        sheet.cell(row=row_, column=1).value = f'Итого по отделению - {count}'
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=3).value = f'Умерших - {ym_count}'
        sheet.cell(row=row_, column=3).font = font
        sheet.row_dimensions[row_].height = 20
        row_ += 1
        sheet.cell(row=row_, column=1).value = 'Из них:'
        sheet.cell(row=row_, column=1).font = font
        sheet.row_dimensions[row_].height = 20
        for dat in data:
            row_ += 1
            sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:C{sheet.cell(row=row_, column=3).row}")
            sheet.cell(row=row_, column=1).value = f'{dat[0]} - {len(dat[1])}'
            sheet.cell(row=row_, column=1).font = font
            sheet.row_dimensions[row_].height = 20
        return sheet
def insert_sheet_P2(**kwargs):
    sheet = kwargs['sheet']
    # data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = list()
    for n,p in enumerate(kwargs['data'],1):
        temp = OrderedDict()
        temp['n'] = n
        temp['fam'] = p.patient.fam
        temp['nib'] = p.sluchay.nib
        temp['im'] = p.patient.im
        temp['ot'] = p.patient.ot
        temp['datr'] = p.patient.datr
        temp['age'] = datetime.now().year - p.patient.datr.year
        temp['nvs'] = p.patient.nvs
        temp['datp'] = p.sluchay.datp
        temp['tm_otd'] = p.sluchay.tm_otd
        temp['datv'] = p.sluchay.datv
        temp['adr'] = p.patient.m_roj
        temp['rab'] = p.patient.rab
        temp['lpy'] = p.sluchay.lpy.naim[:20] if p.sluchay.lpy else None
        temp['otd'] = p.sluchay.otd.naim if p.sluchay.otd else None
        temp['dskz'] = f'{p.sluchay.dskz.kod}-{p.sluchay.dskz.naim}' if p.sluchay.dskz else None
        temp['ds_osl'] = f'{p.sluchay.ds_osl.kod}-{p.sluchay.ds_osl.naim}' if p.sluchay.ds_osl else None
        temp['dsc'] = f'{p.sluchay.dsc.kod}-{p.sluchay.dsc.naim}' if p.sluchay.dsc else None
        temp['dson'] = f'{p.sluchay.dson.kod}-{p.sluchay.dson.naim}' if p.sluchay.dson else None
        temp['vra'] = f'{p.sluchay.le_vr.kod.naim} {p.sluchay.le_vr.kod.ini}' if p.sluchay.le_vr and p.sluchay.le_vr.kod else None
        temp['nib'] = f'N МЭК - {p.sluchay.nib}'
        temp['icx'] = p.sluchay.icx.iz_name if p.sluchay.icx else None
        temp['rslt'] = p.sluchay.rslt.tip_name if p.sluchay.rslt else None
        temp['goc'] = p.sluchay.goc.tip_name if p.sluchay.goc else None
        temp['vds'] = p.sluchay.vds.vds.naim if p.sluchay.vds and p.sluchay.vds.vds else None
        temp['ctkom'] = p.sluchay.vds.ctkom.naim if p.sluchay.vds and p.sluchay.vds.ctkom else None
        temp['vrez'] = p.sluchay.vrez.naim if p.sluchay.vrez else None
        if p.sluchay.wskr:
            temp['wskr'] = p.sluchay.wskr_display()
        else:
            temp['wskr'] = None

        opers = list()
        if p.sluchay.oper.count() > 0:
            for o in p.sluchay.oper.all():
                opers.append(o)
        temp['oper'] = opers
        data.append(temp)
        
    data = sorted(data,key=itemgetter('fam'))
    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row_ = 7
    font = styles.Font(size=15, name='Arial')
    for n, d in enumerate(data, 1):
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=1).value = n
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=2).value = d['fam']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=3).value = f"{d['age']} {d['nvs']}"
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=4).value = f"{d['datp'].strftime('%d.%m.%Y')}" if d['datp'] else None
        sheet.cell(row=row_, column=4).font = font
        sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=5).value = f"{d['datv'].strftime('%d.%m.%Y')}" if d['datv'] else None
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:F{sheet.cell(row=row_ + 2, column=6).row}")
        sheet.cell(row=row_, column=6).value = d['adr']
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_, column=6).font = font
        sheet.merge_cells(f"G{sheet.cell(row=row_, column=7).row}:G{sheet.cell(row=row_ + 1, column=7).row}")
        sheet.cell(row=row_, column=7).value = d['otd']
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        row_ += 3
        sheet.row_dimensions[row_-2].height = 20
        sheet.cell(row=row_ - 2, column=2).value = f"{d['im']} {d['ot']}"
        sheet.cell(row=row_ - 2, column=2).font = font
        sheet.cell(row=row_ - 2, column=2).alignment = styles.Alignment(vertical="center")
        sheet.cell(row=row_ - 2, column=3).value = f"{d['datr'].strftime('%d.%m.%Y')}" if d['datr'] else None
        sheet.cell(row=row_ - 2, column=3).font = font
        sheet.cell(row=row_ - 2, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_ - 2, column=4).value = d['tm_otd']
        sheet.cell(row=row_ - 2, column=4).font = font
        sheet.cell(row=row_ - 2, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=6).value = d['rab']
        sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
        sheet.row_dimensions[row_-1].height = 20
        sheet.cell(row=row_ - 1, column=7).value = d['vra']
        sheet.cell(row=row_ - 1, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_ - 1, column=7).font = font
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=2).value = d['nib']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(vertical="center")
        sheet.merge_cells(f"C{sheet.cell(row=row_, column=3).row}:D{sheet.cell(row=row_, column=4).row}")
        sheet.cell(row=row_, column=3).value = d['icx']
        sheet.cell(row=row_, column=3).font = font
        sheet.merge_cells(f"E{sheet.cell(row=row_, column=5).row}:F{sheet.cell(row=row_, column=6).row}")
        sheet.cell(row=row_, column=5).value = d['rslt']
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(vertical="center")
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=7).value = d['wskr']
        sheet.cell(row=row_, column=7).font = font
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=2).value = d['vrez']
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(vertical="center")
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=3).value = d['goc']
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(vertical="center")
        sheet.merge_cells(f"D{sheet.cell(row=row_, column=4).row}:F{sheet.cell(row=row_, column=6).row}")
        sheet.cell(row=row_, column=4).value = d['lpy']
        sheet.cell(row=row_, column=4).font = font
        row_ += 2
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
        sheet.merge_cells(f"A{sheet.cell(row=row_ - 1, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
        sheet.row_dimensions[row_-1].height = 25
        sheet.cell(row=row_ - 1, column=1).value = d['vds']
        sheet.cell(row=row_ - 1, column=1).font = font
        sheet.cell(row=row_ - 1, column=1).alignment = styles.Alignment(vertical="center", wrap_text=True)
        sheet.merge_cells(f"C{sheet.cell(row=row_ - 1, column=3).row}:E{sheet.cell(row=row_, column=5).row}")
        sheet.merge_cells(f"C{sheet.cell(row=row_, column=3).row}:E{sheet.cell(row=row_, column=5).row}")
        sheet.cell(row=row_ - 1, column=3).value = d['ctkom']
        sheet.cell(row=row_ - 1, column=3).alignment = styles.Alignment(vertical="center", wrap_text=True)
        sheet.cell(row=row_ - 1, column=3).font = font
        if d['dskz']:
            sheet.merge_cells(f"F{sheet.cell(row=row_ - 1, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_ - 1].height = 20
            sheet.cell(row=row_ - 1, column=6).value = d['dskz']
            sheet.cell(row=row_ - 1, column=6).alignment = styles.Alignment(vertical="center", wrap_text=True)
            sheet.cell(row=row_ - 1, column=6).font = font
        if d['ds_osl']:
            row_ += 1
            sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=6).value = d['ds_osl']
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
            sheet.cell(row=row_, column=6).font = font
        if d['dsc']:
            row_ += 1
            sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=6).value = d['dsc']
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
            sheet.cell(row=row_, column=6).font = font
        if d['dson']:
            row_ += 1
            sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=6).value = d['dson']
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
            sheet.cell(row=row_, column=6).font = font
        if len(d['oper']) > 0:
            for o in d['oper']:
                row_ += 1
                sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
                sheet.row_dimensions[row_].height = 20
                sheet.cell(row=row_,column=6).value = f"{o.dato.strftime('%d.%m.%Y') if o.dato else None} {o.kod_op.kod if o.kod_op else ''} {o.kod_op.naim if o.kod_op else ''}"
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
                sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=1).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=2).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=3).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=4).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=5).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=6).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=7).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
def insert_sheet_P19(**kwargs):
    sheet = kwargs['sheet']
    # data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = []
    for p in kwargs['data']:
        ord = OrderedDict()
        ord['fio'] = f'{p.patient.fam} {p.patient.im} {p.patient.ot}'
        ord['datr'] = p.patient.datr.strftime('%d.%m.%Y') if p.patient.datr else None
        ord['adr'] = p.patient.m_roj
        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
        ord['goc'] = p.sluchay.goc.tip_name if p.sluchay.goc else None
        ord['pri'] = p.sluchay.pri.naim if p.sluchay.pri else None
        ord['lpy'] = p.sluchay.lpy.naim[:20] if p.sluchay.lpy else None
        data.append(ord)

    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row_ = 7
    font = styles.Font(size=14, name='Arial')
    for n, d in enumerate(data, 1):
        row_ += 1
        sheet.row_dimensions[row_].height = 65
        sheet.cell(row=row_, column=1).value = n
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=1).border = border
        sheet.cell(row=row_, column=2).value = d['fio']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="left",vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=2).border = border
        sheet.cell(row=row_, column=3).value = d['datr']
        sheet.cell(row=row_, column=3).border = border
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=4).value = d['adr']
        sheet.cell(row=row_, column=4).font = font
        sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="left",wrap_text=True)
        sheet.cell(row=row_, column=4).border = border
        sheet.cell(row=row_, column=5).value = d['datp']
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=5).border = border
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=6).value = d['goc']
        sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=6).border = border
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=7).value = d['pri']
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).border = border
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=8).value = d['lpy']
        sheet.cell(row=row_, column=8).border = border
        sheet.cell(row=row_, column=8).font = font
        sheet.cell(row=row_, column=8).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
def insert_sheet_P22(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    otd = kwargs['otd']
    data = []
    otd_list = []
    for p in kwargs['data']:
        ord = OrderedDict()
        ord['fio_nib'] = f'{p.patient.fam} {p.patient.im} {p.patient.ot} {p.sluchay.nib}'
        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
        if p.sluchay.vds:
            if p.sluchay.vds.vds and p.sluchay.vds.vds.kod == 'Д':
                ord['vds'] = 'ВМП БП'
            elif p.sluchay.vds.vds and p.sluchay.vds.vds.kod == '5':
                ord['vds'] = 'ВМП СБ'
            else:
                ord['vds'] = None
        else:
            ord['vds'] = None
        ord['vra_fio'] = f'{p.sluchay.le_vr.kod.kod} {p.sluchay.le_vr.kod.naim} {p.sluchay.le_vr.kod.ini}' \
            if p.sluchay.le_vr and p.sluchay.le_vr.kod else None
        if ord['vds'] == 'ВМП БП':
            ord['metod_vid'] = f'1.{p.sluchay.metod_hmp}.{p.sluchay.vid_hmp}'
        elif ord['vds'] == 'ВМП СБ':
            ord['metod_vid'] = f'2.{p.sluchay.metod_hmp}.{p.sluchay.vid_hmp}'
        else:
            ord['metod_vid'] = None
        oper = get_pop_oper(p)
        ord['oper_kod'] = oper.kod_op.kod if oper is not None and oper.kod_op else None
        ord['vra_oper'] = f'{oper.kodx.kod} {oper.kodx.naim} {oper.kodx.ini}' if oper is not None and oper.kodx else None
        ord['otd'] = p.sluchay.otd.naim if p.sluchay.otd else None
        otd_list.append(p.sluchay.otd.naim if p.sluchay.otd else None)
        ord['an'] = f'{oper.kodan.kod} {oper.kodan.naim} {oper.kodan.ini}' if oper is not None and oper.kodan else None
        data.append(ord)
    row_ = 6
    font = styles.Font(size=14, name='Arial')
    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    if otd:
        sheet.cell(row=5, column=1).value = f'Отделение - {otd}'
    
    otd_list.sort()
    for en,o in enumerate(set(otd_list)):
        nn=0
        # if en != 0:
        d = ['N п/п','Фамилия И.О. пациента (N ист.)',' Дата пост.    Дата выбыт.','Источник финанс-я','Таб.N ФИО леча-щего врача',
        'Код вида ВТМП метод','Код операции','Таб.N ФИО Хирурга','Отделение','Таб.N ФИО анестези.']
        if en == 0:
            row_+=1
        else:
            row_ +=3
        sheet.row_dimensions[row_].height = 60
        for n,f in enumerate(d,1):
            sheet.cell(row=row_, column=n).value = f
            sheet.cell(row=row_, column=n).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            sheet.cell(row=row_, column=n).font = font
            sheet.cell(row=row_,column=n).border = border
        else:

            sheet.merge_cells(f"A{sheet.cell(row=row_-1, column=1).row}:J{sheet.cell(row=row_ -1, column=13).row}")
            sheet.cell(row=row_-1, column=1).value = o
            sheet.cell(row=row_-1, column=1).alignment =  styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_-1, column=1).font = font


        for n,d in enumerate(data,1):
            if o == d['otd']:
                nn+=1
                row_ += 1
                sheet.row_dimensions[row_].height = 45
                sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:A{sheet.cell(row=row_ + 1, column=1).row}")
                sheet.cell(row=row_, column=1).value = nn
                sheet.cell(row=row_, column=1).font = font
                sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"B{sheet.cell(row=row_, column=2).row}:B{sheet.cell(row=row_ + 1, column=2).row}")
                sheet.cell(row=row_, column=2).value = d['fio_nib']
                sheet.cell(row=row_, column=2).alignment = styles.Alignment(wrap_text=True)
                sheet.cell(row=row_, column=2).font = font
                sheet.cell(row=row_, column=3).value = d['datp']
                sheet.cell(row=row_, column=3).alignment =  styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=3).font = font
                sheet.cell(row=row_ + 1,column=3).value = d['datv']
                sheet.cell(row=row_ + 1,column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_ + 1, column=3).font = font
                sheet.merge_cells(f"D{sheet.cell(row=row_, column=4).row}:D{sheet.cell(row=row_ + 1, column=4).row}")
                sheet.cell(row=row_, column=4).value = d['vds']
                sheet.cell(row=row_, column=4).font = font
                sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"E{sheet.cell(row=row_, column=5).row}:E{sheet.cell(row=row_ + 1, column=5).row}")
                sheet.cell(row=row_, column=5).value = d['vra_fio']
                sheet.cell(row=row_, column=5).font = font
                sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:F{sheet.cell(row=row_ + 1, column=6).row}")
                sheet.cell(row=row_, column=6).value = d['metod_vid']
                sheet.cell(row=row_, column=6).font = font
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"G{sheet.cell(row=row_, column=7).row}:G{sheet.cell(row=row_ + 1, column=7).row}")
                sheet.cell(row=row_, column=7).value = d['oper_kod']
                sheet.cell(row=row_, column=7).font = font
                sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"H{sheet.cell(row=row_, column=8).row}:H{sheet.cell(row=row_ + 1, column=8).row}")
                sheet.cell(row=row_, column=8).value = d['vra_oper']
                sheet.cell(row=row_, column=8).font = font
                sheet.cell(row=row_, column=8).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                sheet.merge_cells(f"I{sheet.cell(row=row_, column=9).row}:I{sheet.cell(row=row_ + 1, column=9).row}")
                sheet.cell(row=row_, column=9).font = font
                sheet.cell(row=row_, column=9).value = d['otd']
                sheet.cell(row=row_, column=9).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
                sheet.merge_cells(f"J{sheet.cell(row=row_, column=10).row}:J{sheet.cell(row=row_ + 1, column=10).row}")
                sheet.cell(row=row_, column=10).font = font
                sheet.cell(row=row_, column=10).value = d['an']
                sheet.cell(row=row_, column=10).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                for c in range(1,11):
                    sheet.cell(row=row_+1, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                row_+=1
def insert_sheet_P24(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    #kwargs['data']
    data = []
    for n,p in enumerate(kwargs['data'],1):
        ord = OrderedDict()
        ord['n'] = n
        ord['nib'] = p.sluchay.nib
        ord['fio'] = f'{p.patient.fam} {p.patient.im} {p.patient.ot}'
        ord['datr'] = p.patient.datr.strftime('%d.%m.%Y') if p.patient.datr else None
        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
        ord['icx'] = p.sluchay.icx.iz_name if p.sluchay.icx else None
        ord['adr'] = p.patient.m_roj
        data.append(ord)
    row = 7
    font = styles.Font(size=14, name='Arial')
    for d in data:
        row+=1
        sheet.row_dimensions[row].height = 33
        for n,i in enumerate(d.values(),1):
            sheet.cell(row=row, column=n).value = i
            if n == 1:
                sheet.cell(row=row, column=n).alignment = styles.Alignment(horizontal="center", vertical="center")
            elif n == 7:
                sheet.cell(row=row, column=n).alignment = styles.Alignment(horizontal="left",vertical="center",wrap_text=True)
            else:
                sheet.cell(row=row, column=n).alignment = styles.Alignment(horizontal="left", vertical="center")
            sheet.cell(row=row, column=n).font = font
            sheet.cell(row=row, column=n).border = border
def insert_sheet_nnn(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = []
    ym = 0
    s1 = 0
    for p in kwargs['data']:
        ord = OrderedDict()
        if p.sluchay.otd_y and p.sluchay.otd_y.naim in ['АРО N1','АРО N2','АРО N3 (ЛДО)','ПРИЕМНОЕ']:
            ym +=1
        if p.sluchay.icx.id_iz == 106:
            s1 +=1
        ord['fio'] = f'{p.sluchay.nib} {p.patient.fam} {p.patient.im[0] if len(p.patient.im) > 0 else ""}.{p.patient.ot[0] if len(p.patient.ot) > 0 else ""}'
        # ord['prof_k'] = p.sluchay.le_vr.prof_k.k_prname if p.sluchay.le_vr and p.sluchay.le_vr.prof_k else None
        ord['prof_k'] = p.sluchay.otd.naim if p.sluchay.otd else None
        ord['age'] = f'{datetime.now().year - p.patient.datr.year} {p.patient.nvs}'
        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
        if p.sluchay.goc and p.sluchay.goc.tip_name == 'Экстренная':
            ord['goc'] = 'экс'
        elif p.sluchay.goc and p.sluchay.goc.tip_name == 'Плановая':
            ord['goc'] = 'пл'
        else:
            ord['goc'] = ''
        ord['vra'] = p.sluchay.le_vr.kod.naim if p.sluchay.le_vr and p.sluchay.le_vr.kod else None
        ord['dskz'] = f'{p.sluchay.dskz.kod}-{p.sluchay.dskz.naim}' if p.sluchay.dskz else ''
        ord['dspat'] = f'{p.sluchay.dspat.kod}-{p.sluchay.dspat.naim}' if p.sluchay.dspat else ''
        ord['pri'] = p.sluchay.pri.naim if p.sluchay.pri else ''
        opers = list()
        if p.sluchay.oper.count() > 0:
            for o in p.sluchay.oper.all().order_by('-pop'):
                opers.append(o)
        ord['oper'] = opers
        # oper = get_pop_oper(p)
        # ord['dato_oper'] = oper.dato.strftime('%d.%m.%Y') if oper is not None else None
        # oper['goc'] = p.sluchay.goc.tip_name if p.sluchay.goc else None
        # ord['vra_oper'] = oper.kod.naim if oper is not None and oper.kod else None
        data.append(ord)

    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row_ = 8
    font = styles.Font(size=11, name='Arial')
    for n, d in enumerate(data, 1):
        row_ += 1
        sheet.row_dimensions[row_].height = 50
        sheet.cell(row=row_, column=1).value = f'{n}   {d["fio"]}'
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=2).value = d['age']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=3).value = d['datp']
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=4).value = d['datv']
        sheet.cell(row=row_, column=4).font = font
        sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=5).value = d['goc']
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=6).value = d['vra']
        sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=7).value = f'1.{d["dskz"]}'
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=7).font = font
        row_+=1
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:F{sheet.cell(row=row_, column=6).row}")
        sheet.row_dimensions[row_].height = 50
        sheet.cell(row=row_, column=1).value = d['prof_k']
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="left",vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=7).value = f'2.{d["dspat"]}'
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        row_+=1
        sheet.cell(row=row_, column=7).value = f'3.{d["pri"]}'
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.row_dimensions[row_].height = 35
        if len(d['oper']) > 0:
            for o in d['oper']:
                row_+=1
                sheet.cell(row=row_,column=4).value = o.dato.strftime('%d.%m.%Y') if o.dato else ''
                sheet.cell(row=row_, column=4).font = font
                sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="left")
                sheet.cell(row=row_, column=6).value = o.kodx.naim if o.kodx else ''
                sheet.cell(row=row_, column=6).font = font
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="left")
                sheet.cell(row=row_, column=7).value = f'4.{o.kod_op.kod if o.kod_op else ""}{"+" if o.pop == True else ""}'
                sheet.cell(row=row_, column=7).font = font
        for c in range(1,8):
            sheet.cell(row=row_, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

    row_ += 3
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
    sheet.cell(row=row_, column=1).value = f'Умерло всего - {len(data)}'
    sheet.cell(row=row_, column=1).font = font
    row_+=1
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
    sheet.cell(row=row_, column=1).value = f'в том числе'
    sheet.cell(row=row_, column=1).font = font
    row_ += 1
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:C{sheet.cell(row=row_, column=3).row}")
    sheet.cell(row=row_, column=1).value = f'в реанимации и прием.отд-ии - {ym}'
    sheet.cell(row=row_, column=1).font = font
    row_ += 1
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
    sheet.cell(row=row_, column=1).value = f'в 1-е сутки - {s1}'
    sheet.cell(row=row_, column=1).font = font
def insert_sheet_a_oth_5(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    user = kwargs['user']

    data = get_list_otd(kwargs['data'])
    data_filter = get_list_otd(kwargs['data_filter'])

    row = 8
    font = styles.Font(size=14, name='Arial')
    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()

    all_sl = []
    all_sl_filter = []
    for d in data:
        row+=1
        rez_data = get_rez_a_oth_5(d[1])
        all_sl.append(rez_data) 
        sheet.cell(row=row, column=1).value = d[0][0]
        fields_data = [i for i in range(1,15) if i % 2 == 1]
        for f in range(len(fields_data)):
            sheet.cell(row=row, column=1+fields_data[f]).value = rez_data[f] if rez_data[f] != 0 else None
            sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
        if len(data_filter) > 0:
            for df in data_filter:
                if d[0] == df[0]:
                    rez_filter = get_rez_a_oth_5(df[1])
                    all_sl_filter.append(rez_filter) 
                    fields_data = [i for i in range(1,15) if i % 2 == 0]
                    for f in range(len(fields_data)):
                        sheet.cell(row=row, column=1+fields_data[f]).value = rez_filter[f] if rez_filter[f] != 0 else None
                        sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
    else:
        age_data = 0
        age_data_filter = 0
        row+=1
        sheet.cell(row=row, column=1).value = 'ВСЕГО'
        r = None
        for o in range(len(all_sl)):
            if o == 0:
                r = numpy.array(all_sl[o])
            else:
                r+= numpy.array(all_sl[o])
        rez_data = r.tolist()
        fields_data = [i for i in range(1,15) if i % 2 == 1]
        age_data = float('{0:.2f}'.format(rez_data[7]/rez_data[0]))
        for f in range(len(fields_data)):
            sheet.cell(row=row, column=1+fields_data[f]).value = rez_data[f] if rez_data[f] != 0 else None
            sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
        if len(all_sl_filter) != 0:
            r = None
            for o in range(len(all_sl_filter)):
                if o == 0:
                    r = numpy.array(all_sl_filter[o])
                else:
                    r+= numpy.array(all_sl_filter[o])
            rez_filter = r.tolist()
            age_data_filter = float('{0:.2f}'.format(rez_filter[7]/rez_filter[0]))
            fields_data = [i for i in range(1,15) if i % 2 == 0]
            for f in range(len(fields_data)):
                sheet.cell(row=row, column=1+fields_data[f]).value = rez_filter[f] if rez_filter[f] != 0 else None
                sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
        row+=1
        sheet.cell(row=row, column=1).value = 'Средний возраст (кол-во лет)'
        sheet.cell(row=row, column=2).value = age_data if age_data != 0 else None
        sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=3).value = age_data_filter if age_data_filter != 0 else None
        sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")



def get_rez_a_oth_5(data):
    bf = BetterFilter()
    sp = CountSluchaySpecification() ^ GocEkSpecification() ^ ProfKNSpecification() ^ \
         OperCountSpecification() ^ RezUmerSpecification() ^ PoslOperKdSpecification() ^ AgeSpecification()
    all_temp = []
    for patient in data:
        for p in bf.filter(patient,sp):
            temp = bf.format_list(p)
            for t in range(len(temp)):
                if temp[t] == 'None':
                    temp[t] = 0
            all_temp.append([int(i) for i in temp])
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(7)]
    all_temp.insert(3,0)
    all_temp[3] = float('{0:.2f}'.format(all_temp[2]/all_temp[0]))
    return all_temp

def insert_sheet_a_oth_7(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    user = kwargs['user']
    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=6, column=1).value = str(name).capitalize()    
    data = get_list_otd_prof(kwargs['data'])

    row = 9
    font = styles.Font(size=10, name='Arial')
    fs = ['Выбыло','Койко/день','Сред.к/день','Оперировано','К-во операц','Умерло']
    #Всего пациентов
    sl_all = []
    for d in data:
        for b in range(1,13):
            sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        row+=1
        sheet.row_dimensions[row].height = 25
        sheet.cell(row=row, column=1).value = d[0][0]
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        #Список итог всех пациентов в отделении
        otd_prof_sl_all = []
        for prof in d[1]:
            sheet.cell(row=row, column=2).value = prof[0]
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = get_rez_7(prof[1])
            otd_prof_sl_all.append(rez)
            sl_all.append(rez)
            for f in range(len(fs)):
                sheet.cell(row=row, column=3).value = fs[f]
                for b in range(3,13):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                for v in range(len(rez[f])):
                    sheet.cell(row=row, column=4+v).value = rez[f][v] if rez[f][v] != 0 else None
                    sheet.cell(row=row, column=4+v).alignment = styles.Alignment(horizontal="center", vertical="center")
                row+=1
            else:
                row-=1
                for b in range(2,13):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            if len(d[1]) > 1:
                row+=1
        else:
            if len(d[1]) > 1:
                r = None
                for o in range(len(otd_prof_sl_all)):
                    if o == 0:
                        r = numpy.array(otd_prof_sl_all[o])
                    else:
                        r+= numpy.array(otd_prof_sl_all[o])
                rez = r.tolist()
                rez[2] = (numpy.array(rez[1])/numpy.array(rez[0])).tolist()
                rez[2][1] = rez[2][2] + rez[2][3] + rez[2][4] 

                sheet.cell(row=row, column=2).value = 'Итого'
                for f in range(len(fs)):
                    sheet.cell(row=row, column=3).value = fs[f]
                    for b in range(3,13):
                        sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                    for v in range(len(rez[f])):
                        sheet.cell(row=row, column=4+v).value =  float('{0:.2f}'.format(rez[f][v])) if rez[f][v] != None and rez[f][v] != 0 else None
                        sheet.cell(row=row, column=4+v).alignment = styles.Alignment(horizontal="center", vertical="center")
                    row+=1
                else:
                    row-=1
                    for b in range(2,13):
                        sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
    else:
        row+=1
        for b in range(1,13):
            sheet.cell(row=row, column=b).border = styles.Border(top=styles.Side(border_style='thin', color='000000'))
        r = None
        for o in range(len(sl_all)):
            if o == 0:
                r = numpy.array(sl_all[o])
            else:
                r+= numpy.array(sl_all[o])
        rez = r.tolist()
        # rez = rez[0]
        rez[2] = (numpy.array(rez[1])/numpy.array(rez[0])).tolist()
        rez[2][1] = rez[2][2] + rez[2][3] + rez[2][4] 

        sheet.cell(row=row, column=1).value = 'Всего'
        for f in range(len(fs)):
                sheet.cell(row=row, column=3).value = fs[f]
                for b in range(3,13):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                for v in range(len(rez[f])):
                    sheet.cell(row=row, column=4+v).value =  float('{0:.2f}'.format(rez[f][v])) if rez[f][v] != None and rez[f][v] != 0 else None
                    sheet.cell(row=row, column=4+v).alignment = styles.Alignment(horizontal="center", vertical="center")
                row+=1

def get_rez_7(data):
    bf = BetterFilter()
    all_sp = IsfinKastSpecification('all') ^ IsfinKastSpecification('all','tym_oms') ^ IsfinKastSpecification('all','bez_polis') ^ IsfinKastSpecification('all','dr_oms') ^ \
        IsfinKastSpecification('all','vmp_sb') ^ IsfinKastSpecification('all','vmp_sv') ^ IsfinKastSpecification('all','dms') ^ IsfinKastSpecification('all','fss') 
    all_temp = []
    for patient in data:
        for p in bf.filter(patient,all_sp):
            all_temp.append([int(i) for i in bf.format_list(p)])
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(8)]

    kd_sp = IsfinKastSpecification('kd') ^ IsfinKastSpecification('kd','tym_oms') ^ IsfinKastSpecification('kd','bez_polis') ^ IsfinKastSpecification('kd','dr_oms') ^ \
        IsfinKastSpecification('kd','vmp_sb') ^ IsfinKastSpecification('kd','vmp_sv') ^ IsfinKastSpecification('kd','dms') ^ IsfinKastSpecification('kd','fss') 
    kd_temp = []
    for patient in data:
        for p in bf.filter(patient,kd_sp):
            kd_temp.append([int(i) for i in bf.format_list(p)])
    kd_temp = [sum([kd_temp[i][x] for i in range(len(kd_temp))]) for x in range(8)]

    oper_sp = IsfinKastSpecification('oper') ^ IsfinKastSpecification('oper','tym_oms') ^ IsfinKastSpecification('oper','bez_polis') ^ IsfinKastSpecification('oper','dr_oms') ^ \
        IsfinKastSpecification('oper','vmp_sb') ^ IsfinKastSpecification('oper','vmp_sv') ^ IsfinKastSpecification('oper','dms') ^ IsfinKastSpecification('oper','fss') 
    oper_temp = []
    for patient in data:
        for p in bf.filter(patient,oper_sp):
            oper_temp.append([int(i) for i in bf.format_list(p)])
    oper_temp = [sum([oper_temp[i][x] for i in range(len(oper_temp))]) for x in range(8)]

    oper_count_sp = IsfinKastSpecification('oper_count') ^ IsfinKastSpecification('oper_count','tym_oms') ^ IsfinKastSpecification('oper_count','bez_polis') ^ IsfinKastSpecification('oper_count','dr_oms') ^ \
        IsfinKastSpecification('oper_count','vmp_sb') ^ IsfinKastSpecification('oper_count','vmp_sv') ^ IsfinKastSpecification('oper_count','dms') ^ IsfinKastSpecification('oper_count','fss') 
    oper_count_temp = []
    for patient in data:
        for p in bf.filter(patient,oper_count_sp):
            oper_count_temp.append([int(i) for i in bf.format_list(p)])
    oper_count_temp = [sum([oper_count_temp[i][x] for i in range(len(oper_count_temp))]) for x in range(8)]

    ymer_sp = IsfinKastSpecification('ymer') ^ IsfinKastSpecification('ymer','tym_oms') ^ IsfinKastSpecification('ymer','bez_polis') ^ IsfinKastSpecification('ymer','dr_oms') ^ \
        IsfinKastSpecification('ymer','vmp_sb') ^ IsfinKastSpecification('ymer','vmp_sv') ^ IsfinKastSpecification('ymer','dms') ^ IsfinKastSpecification('ymer','fss') 
    ymer_temp = []
    for patient in data:
        for p in bf.filter(patient,ymer_sp):
            ymer_temp.append([int(i) for i in bf.format_list(p)])
    ymer_temp = [sum([ymer_temp[i][x] for i in range(len(ymer_temp))]) for x in range(8)]
    medium_kd = []
    for k in range(len(kd_temp)):
        try:
            medium_kd.append(float('{0:.2f}'.format(kd_temp[k]/all_temp[k])))
        except ZeroDivisionError:
            medium_kd.append(0)

    # all_temp.insert(1,all_temp[2]+all_temp[3]+all_temp[4])
    # kd_temp.insert(1,kd_temp[2]+kd_temp[3]+kd_temp[4])
    # medium_kd.insert(1,medium_kd[2]+medium_kd[3]+medium_kd[4])
    # oper_temp.insert(1,oper_temp[2]+oper_temp[3]+oper_temp[4])
    # oper_count_temp.insert(1,oper_count_temp[2]+oper_count_temp[3]+oper_count_temp[4])
    # ymer_temp.insert(1,ymer_temp[2]+ymer_temp[3]+ymer_temp[4])

    all_temp.insert(1,0)
    kd_temp.insert(1,0)
    medium_kd.insert(1,0)
    oper_temp.insert(1,0)
    oper_count_temp.insert(1,0)
    ymer_temp.insert(1,0)

    all_temp[1] = all_temp[2]+all_temp[3]+all_temp[4]
    kd_temp[1] = kd_temp[2]+kd_temp[3]+kd_temp[4]
    medium_kd[1] = medium_kd[2]+medium_kd[3]+medium_kd[4]
    oper_temp[1] = oper_temp[2]+oper_temp[3]+oper_temp[4]
    oper_count_temp[1] = oper_count_temp[2]+oper_count_temp[3]+oper_count_temp[4]
    ymer_temp[1] = ymer_temp[2]+ymer_temp[3]+ymer_temp[4]
    


    return [all_temp,kd_temp,medium_kd,oper_temp,oper_count_temp,ymer_temp]



def insert_sheet_a_oth_20(**kwargs):
    sheet = kwargs['sheet']
    # data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']

    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()

    bf = BetterFilter()
    sp = OtdSpecification() ^ ProfkSpecification() ^ OperCountSpecification() ^ EndosOper() ^ OsloCountAllSpecification()

    temp = []
    for d in kwargs['data']:
        for p in bf.filter(d, sp):
            temp.append(bf.format_list(p))
            
    res_set = list(set([t[0] for t in temp]))
    res = []
    for r in res_set:
        res.append([
            r,
            []
        ])
    for r in res:
        for t in temp:
            if r[0] == t[0]:
                r[1].append(t[1])
            r[1] = list(set(r[1]))

    res_set.clear()
    for r in res:
        tt = []
        tt.append(r[0])
        for rr in r[1]:
            tt.append([rr,0,0,0,0])
        res_set.append(tt)
    # print(res_set)
    # for t in temp:
    #     x = (t[0],(t[1],None,None,None,None))
    #     res_set.add(x)
    # res_set = [list(r[rr]) for r in res_set for rr in range(len(r)) if r!=0 ]

    # print(res_set)
    # for t in temp:
    #     for r in res_set:
    #         if r[0] == t[0] and r[1] == t[1]:
    #             pass
    #             r[2] = r[2] + t[2] if r[2] != None else t[2]
    #             r[3] = r[3] + t[3] if r[3] != None else t[3]
    #             r[4] = r[4] + t[4] if r[4] != None else t[4]
    row = 7

    for i, res in enumerate(res_set):
        row +=1
        for l, r in enumerate(res):
            if l == 0:
                sheet.cell(row=row, column=1 + l).value = r if r != 0 else ''
            else:
                for n,prof in enumerate(r):
                    if n == 0:
                        sheet.cell(row=row, column=2).value = r[0] if r[0] != 0 else ''
                    else:
                        row += 1
                        sheet.cell(row=row, column=2).value = r[0] if r[0] != 0 else ''


def get_rez_32_list(data,ds):
    temp = []
    ds_list = []
    ds_list_T51 = ['T51','T51.0','T51.1','T51.2','T51.3','T51.4','T51.5','T51.6','T51.7','T51.8','T51.9']
    ds_list_T40 = ['T40','T40.0','T40.1','T40.2','T40.3','T40.4','T40.5','T40.6','T40.7','T40.8','T40.9']
    ds_list_T40_6 = ['T40.6']
    ds_list_T43 = ['T43','T43.0','T43.1','T43.2','T43.3','T43.4','T43.5','T43.6','T43.7','T43.8','T43.9']
    ds_list_T39 = ['T39','T39.0','T39.1','T39.2','T39.3','T39.4','T39.5','T39.6','T39.7','T39.8','T39.9']
    not_ds = ds_list_T51+ds_list_T40+ds_list_T40_6+ds_list_T43+ds_list_T39

    if ds == 'T51':
        ds_list = ds_list_T51
    elif ds == 'T40':
        ds_list = ds_list_T40
    elif ds == 'T40_6':
        ds_list = ds_list_T40_6
    elif ds == 'T43':
        ds_list = ds_list_T43
    elif ds == 'T39':
        ds_list = ds_list_T39
    

    for d in data:
        if ds != '':
            if d.sluchay.dskz and d.sluchay.dskz.kod in ds_list:
                temp.append(rez_32_list_ord(d))
        else:
            if d.sluchay.dskz and d.sluchay.dskz.kod not in not_ds:
             temp.append(rez_32_list_ord(d))


    return temp
def set_rez_32_list(sheet,row,n,d):
    sheet.row_dimensions[row].height = 35
    sheet.cell(row=row, column=1).value = n
    sheet.cell(row=row, column=1).border = border

    sheet.cell(row=row, column=2).value = d['fio']
    sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
    sheet.cell(row=row, column=2).border = border

    sheet.cell(row=row, column=3).value = d['m']
    sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=3).border = border

    sheet.cell(row=row, column=4).value = d['datr']
    sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=4).border = border

    sheet.cell(row=row, column=5).value = d['datp']
    sheet.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=5).border = border

    sheet.cell(row=row, column=6).value = d['datv']
    sheet.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=6).border = border

    sheet.cell(row=row, column=7).value = d['dskz']
    sheet.cell(row=row, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=7).border = border

    sheet.cell(row=row, column=8).value = d['ds_osl']
    sheet.cell(row=row, column=8).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=8).border = border

    sheet.cell(row=row, column=9).value = d['isx']
    sheet.cell(row=row, column=9).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=9).border = border

    sheet.cell(row=row, column=10).value = d['adr_1']
    sheet.cell(row=row, column=10).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
    sheet.cell(row=row, column=10).border = border
    
    sheet.cell(row=row, column=11).value = d['adr_2']
    sheet.cell(row=row, column=11).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
    sheet.cell(row=row, column=11).border = border

def rez_32_list_ord(d):
    ord = OrderedDict()
    f = d.patient.fam
    i = d.patient.im if len(d.patient.im) > 0 else ''
    ot = d.patient.ot if len(d.patient.ot) > 0 else ''
    ord['fio'] = f'{f} {i}.{ot}'
    ord['m'] = d.patient.pol.polname[0] if d.patient.pol else None
    ord['datr'] = d.patient.datr.strftime('%d.%m.%Y') if d.patient.datr else None
    ord['datp'] = d.sluchay.datp.strftime('%d.%m.%Y') if d.sluchay.datp else None
    ord['datv'] = d.sluchay.datv.strftime('%d.%m.%Y') if d.sluchay.datv else None
    ord['dskz'] = d.sluchay.dskz.kod if d.sluchay.dskz else None
    ord['ds_osl'] = d.sluchay.ds_osl.kod if d.sluchay.ds_osl else None
    if d.sluchay.icx:
        if d.sluchay.icx.id_iz == 101:
            ord['isx'] = 'выздор'
        elif d.sluchay.icx.id_iz == 102:
            ord['isx'] = 'улучшен'
        elif d.sluchay.icx.id_iz == 103:
            ord['isx'] = 'без.пер'
        elif d.sluchay.icx.id_iz == 104:
            ord['isx'] = 'ухудшен'
        elif d.sluchay.icx.id_iz == 105:
            ord['isx'] = 'умер'
        elif d.sluchay.icx.id_iz == 106:
            ord['isx'] = 'ум.1-е.сут.'
    else:
        ord['isx'] = ''
    # if 'Тюмень' in d.patient.m_roj:
    if 'р-н' in d.patient.m_roj:
        ord['adr_1'] = 'Тюм.р-н'
    else:
        ord['adr_1'] = 'г.Тюмень'
    # else:
    #     ord['adr_1'] = ''
    ord['adr_2'] = d.patient.m_roj

    return ord

def get_rez_32_info(data,ds):
    ds_list_T51 = ['T51','T51.0','T51.1','T51.2','T51.3','T51.4','T51.5','T51.6','T51.7','T51.8','T51.9']
    ds_list_T40 = ['T40','T40.0','T40.1','T40.2','T40.3','T40.4','T40.5','T40.6','T40.7','T40.8','T40.9']
    ds_list_T40_6 = ['T40.6']
    ds_list_T40_7 = ['T40.7']
    ds_list_T43 = ['T43','T43.0','T43.1','T43.2','T43.3','T43.4','T43.5','T43.6','T43.7','T43.8','T43.9']
    ds_list_T39 = ['T39','T39.0','T39.1','T39.2','T39.3','T39.4','T39.5','T39.6','T39.7','T39.8','T39.9']
    not_ds = ds_list_T51+ds_list_T40+ds_list_T40_6+ds_list_T43+ds_list_T39+ds_list_T40_7

    ds_list = []
    if ds == 'T51':
        ds_list = ds_list_T51
    elif ds == 'T40':
        ds_list = ds_list_T40
    elif ds == 'T40.6':
        ds_list = ds_list_T40_6
    elif ds == 'T40.7':
        ds_list = ds_list_T40_7
    elif ds == 'T43':
        ds_list = ds_list_T43
    elif ds == 'T39':
        ds_list = ds_list_T39
    
    temp1 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    temp2 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    temp3 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

    for d in data:
        if ds != '':
            if d.sluchay.dskz and d.sluchay.dskz.kod in ds_list:
                set_rez_32_info(d,temp1,temp2,temp3)
        else:
            if d.sluchay.dskz and d.sluchay.dskz.kod not in not_ds:
                set_rez_32_info(d,temp1,temp2,temp3)
    
    return [temp1,temp2,temp3]

def set_rez_32_info(d,temp1,temp2,temp3):
    if 'р-н' not in d.patient.m_roj:
        temp1[0]+=1
        if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
            temp2[0]+=1
        if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
            temp3[0]+=1
    if 'р-н' in d.patient.m_roj: 
        temp1[1]+=1
        if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
            temp2[1]+=1
        if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
            temp3[1]+=1

    if d.patient.pol and d.patient.pol.id_pol == 2:
        if 'р-н' not in d.patient.m_roj:
            temp1[2]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[2]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[2]+=1
    if d.patient.pol and d.patient.pol.id_pol == 2:
        if 'р-н' in d.patient.m_roj: 
            temp1[3]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[3]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[3]+=1
    
    if datetime.now().year - d.patient.datr.year < 15:
        if 'р-н' not in d.patient.m_roj:
            temp1[4]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[4]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[4]+=1
    if datetime.now().year - d.patient.datr.year < 15:
        if 'р-н' in d.patient.m_roj: 
            temp1[5]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[5]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[5]+=1
    
    if 15 <= datetime.now().year - d.patient.datr.year <= 17:
        if 'р-н' not in d.patient.m_roj:
            temp1[6]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[6]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[6]+=1
    if 15 <= datetime.now().year - d.patient.datr.year <= 17:
        if 'р-н' in d.patient.m_roj: 
            temp1[7]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[7]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[7]+=1
    
    if 18 <= datetime.now().year - d.patient.datr.year <= 39:
        if 'р-н' not in d.patient.m_roj:
            temp1[8]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[8]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[8]+=1
    if 18 <= datetime.now().year - d.patient.datr.year <= 39:
        if 'р-н' in d.patient.m_roj: 
            temp1[9]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[9]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[9]+=1
    
    if 18 <= datetime.now().year - d.patient.datr.year <= 19:
        if 'р-н' not in d.patient.m_roj:
            temp1[10]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[10]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[10]+=1
    if 18 <= datetime.now().year - d.patient.datr.year <= 19:
        if 'р-н' in d.patient.m_roj: 
            temp1[11]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[11]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[11]+=1
    
    if 20 <= datetime.now().year - d.patient.datr.year <= 39:
        if 'р-н' not in d.patient.m_roj:
            temp1[12]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[12]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[12]+=1
    if 20 <= datetime.now().year - d.patient.datr.year <= 39:
        if 'р-н' in d.patient.m_roj: 
            temp1[13]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[13]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[13]+=1
    
    if datetime.now().year - d.patient.datr.year >= 40:
        if 'р-н' not in d.patient.m_roj:
            temp1[14]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[14]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[14]+=1
    if datetime.now().year - d.patient.datr.year >= 40:
        if 'р-н' in d.patient.m_roj: 
            temp1[15]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[15]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[15]+=1

def insert_sheet_a_oth_32(**kwargs):
    sheet = kwargs['sheet'][0]
    sheet1 = kwargs['sheet'][1]
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    ###Инфо
    rez_T51 = get_rez_32_info(kwargs['data'],'T51')
    rez_T40 = get_rez_32_info(kwargs['data'],'T40')
    rez_T40_7 = get_rez_32_info(kwargs['data'],'T40.7')
    rez_T40_6 = get_rez_32_info(kwargs['data'],'T40.6')
    rez_T43 = get_rez_32_info(kwargs['data'],'T43')
    rez_T39 = get_rez_32_info(kwargs['data'],'T39')
    rez_none = get_rez_32_info(kwargs['data'],'')



    row = 6
    for data in rez_T51:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
   

    for data in rez_T40:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    for data in rez_T40_7:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")


    for data in rez_T40_6:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")

    
    for data in rez_T43:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    for data in rez_T39:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    for data in rez_none:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    ### Список
    rez_T51 = get_rez_32_list(kwargs['data'],'T51')
    rez_T40 = get_rez_32_list(kwargs['data'],'T40')
    rez_T40_6 = get_rez_32_list(kwargs['data'],'T40_6')
    rez_T43 = get_rez_32_list(kwargs['data'],'T43')
    rez_T39 = get_rez_32_list(kwargs['data'],'T39')
    rez_none = get_rez_32_list(kwargs['data'],'')

    row = 3
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления алкоголем и спиртами (Т51-Т51.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T51,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)

    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления наркотическими средствами (Т40-Т40.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T40,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления курительными смесями (Т40.6)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T40_6,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
        
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления психотропными средствами смесями (Т43-Т43.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T43,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления (Т39-Т39.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T39,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Другие отравления и токсич.воздействия веществ'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_none,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)

class GroupP1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
        self.group_st_ot = request.get('group_st_ot')
        self.otdels = json.loads(request.get('otdels'))
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            if self.group_st_ot == 'st':
                otds = list(set([p.sluchay.otd.naim for p in patients.patients if p.sluchay.otd]))
            elif self.group_st_ot == 'ot':
                otds = list(set([p.sluchay.otd.naim for p in patients.patients if p.sluchay.otd and p.sluchay.otd.naim in self.otdels]))
            data = list()
            count = 0
            ym_count = 0

            for o in sorted(otds):
                temp = []
                for p in patients.patients:
                    if p.sluchay.otd and p.sluchay.otd.naim == o:
                        ord = OrderedDict()
                        count += 1
                        ord['nib'] = p.sluchay.nib[4:]
                        f = p.patient.fam
                        i = p.patient.im[0] if len(p.patient.im) > 0 else ''
                        ot = p.patient.ot[0] if len(p.patient.ot) > 0 else ''
                        ord['fio'] = f'{f} {i}.{ot}'
                        ord['datr'] = p.patient.datr.strftime('%d.%m.%Y') if p.patient.datr else None
                        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
                        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
                        ord['isx'] = 'умр.' if p.sluchay.icx and p.sluchay.icx.id_iz in [105, 106] else None
                        if ord['isx']:
                            ym_count += 1
                        ord['dskz'] = f'{p.sluchay.dskz.kod}' if p.sluchay.dskz else None
                        ord['prof_k'] = p.sluchay.le_vr.prof_k.k_prname if p.sluchay.le_vr and p.sluchay.le_vr.prof_k else None
                        ord['vds'] = p.sluchay.vds.vds.naim[:5] if p.sluchay.vds and p.sluchay.vds.vds else None
                        temp.append(ord)
                data.append([o, temp])

            for d in data:
                for n,i in enumerate(d):
                    if n != 0:
                        d[n] = sorted(d[n],key=itemgetter('fio'))

            dic = dict([('sheet',sheet),('name',self.user.statistics_type.name),
                        ('date_1',self.date_1),('date_2',self.date_2),('data',data),
                        ('count',count),('ym_count',ym_count)])
            sheet = insert_sheet_P1(**dic)
            wb.save(self.path() + f'group_p1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download', 'text': self.path() + f'group_p1_{self.user.user.id}.xlsx',
                                                           'name': 'Отчет о выбывших'})
class GroupP2(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user,request)
        self.request = request
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p2_2.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            sheet = wb.active
            patients = PatientsDataFiltrs(self.date_1,self.date_2,self.user,self.request)
            dic = dict([('sheet',sheet),('data',patients.patients),('name',self.user.statistics_type.name),
                        ('date_1',self.date_1),('date_2',self.date_2)])
            insert_sheet_P2(**dic)
            wb.save(self.path() + f'group_p2_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p2_{self.user.user.id}.xlsx',
                                                           'name': 'Отчет о выбывших расшир.'})
        else:
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'text': 'Отсутствует шаблон - group_p1.xlsx'})
            raise FileNotFoundError('Отсутствует шаблон - group_p1.xlsx')
class GroupP3(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP4(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP5(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class GroupP6(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP7(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP8(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP9(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP10(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP11(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP12(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP13(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP14(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP15(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP16(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP17(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP18(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP19(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p19.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if (p.sluchay.icx and p.sluchay.icx.id_iz in(105,106))\
                        or (p.sluchay.rslt and p.sluchay.rslt.id_tip in (105,106)):
                    if (datetime.now().year - p.patient.datr.year) <= 65:
                        ## в первые сутки ???
                        if p.sluchay.datp == p.sluchay.datv:
                            data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_P19(**dic)
            wb.save(self.path() + f'group_p19_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                         {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p19_{self.user.user.id}.xlsx',
                                                           'name': ''})

        else:
            pass
class GroupP20(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP21(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP22(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        filters = json.loads(request.get('filters'))['filter']
        self.otd = filters['otd']['otd'] if filters['otd']['otd'] else None
        ### Уточнить фильтруем по этим полям ?
        self.metod_hmp = filters['metod_hmp']['metod_hmp'] if filters['metod_hmp']['metod_hmp'] else None
        self.vid_hmp = filters['vid_hmp']['vid_hmp'] if filters['vid_hmp']['vid_hmp'] else None
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p22.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsDataFiltrs(self.date_1,self.date_2,self.user,self.request)
            data = []
            if self.otd != None:
                for p in patients.patients:
                    if p.sluchay.otd and p.sluchay.otd.naim == self.otd and p.sluchay.tip_oms == '2':
                        data.append(p)
            else:
                for p in patients.patients:
                    if p.sluchay.tip_oms == '2':
                        data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('otd',self.otd)])
            sheet = insert_sheet_P22(**dic)
            wb.save(self.path() + f'group_p22_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p22_{self.user.user.id}.xlsx'})
class GroupP23(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP24(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p24.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsDataFiltrs(self.date_1, self.date_2, self.user, self.request)
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('user', self.user)])
            insert_sheet_P24(**dic)
            wb.save(self.path() + f'group_p24_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p24_{self.user.user.id}.xlsx',
                                                           'name': ''})
class GroupP25(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass

class VaultOtdTTT(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
    def create(self):
        pass
class VaultOtdNNN(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('vault_otd_rep_nnn.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim in ('НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3'):
                    if (p.sluchay.icx and p.sluchay.icx.id_iz in (105, 106)) \
                            or (p.sluchay.rslt and p.sluchay.rslt.id_tip in (105, 106)):
                        data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_nnn(**dic)
            wb.save(self.path() + f'group_nnn_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_vault_otd',
                                                           'text': self.path() + f'group_nnn_{self.user.user.id}.xlsx'})
class VaultOtdHH(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
        # file = self.is_file('vault_otd_rep_nnn.xlsx')
        # if file:
        #     wb = load_workbook(file)
        #     sheet = wb.active
        #     os.remove(file)
        #     patients = PatientsData(self.date_1, self.date_2, self.user)
        #     patients.sluchays()
        #     data = []
        #     for p in patients.patients:
        #         if p.sluchay.otd and p.sluchay.otd.naim in ['НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3']:
        #             if

class AOth1(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})

class AOth2(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth3(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth4(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth5(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('a_oth_5.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()

            filters = json.loads(self.request.get('filters'))
            if len(filters['filter']) > 0:
                patients_filter = PatientsDataFiltrs(self.date_1, self.date_2, self.user, self.request)
                data_filter = patients_filter.patients
            else:
                data_filter = []
            dic = dict([('sheet', sheet), ('data', patients.patients),('data_filter',data_filter), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('user',self.user)])
            insert_sheet_a_oth_5(**dic)
            wb.save(self.path() + f'a_oth_5_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_5_{self.user.user.id}.xlsx',
                                                           'btn':'download_a_oth_5'
                                                           },
                                                          )
class AOth6(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth7(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_7.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('user',self.user)])
            insert_sheet_a_oth_7(**dic)
            wb.save(self.path() + f'a_oth_7_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_7_{self.user.user.id}.xlsx',
                                                           'btn':'download_a_oth_7'
                                                           },
                                                          )


class AOth8(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth9(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth10(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth11(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth12(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth13(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth14(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth15(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth16(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth17(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth18(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth19(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth20(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_20.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            sheet = wb.active
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_a_oth_20(**dic)
            wb.save(self.path() + f'a_oth_20_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'a_oth_20_{self.user.user.id}.xlsx'})
class AOth21(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth22(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth23(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth24(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth25(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth26(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth27(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth28(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth29(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth30(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth31(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass
class AOth32(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_32.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            # sheet = wb.active
            sheet=wb.get_sheet_by_name('инфо')
            sheet1=wb.get_sheet_by_name('список')
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'ТОКСИКОЛОГИЯ':
                    data.append(p) 

            dic = dict([('sheet', [sheet,sheet1]), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])

            insert_sheet_a_oth_32(**dic)
            wb.save(self.path() + f'a_oth_32_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'a_oth_32_{self.user.user.id}.xlsx'})

class AOth33(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass





















def ReferenceReport(user,request):
    print(request)
    type_fun = request.get('group_p_list')
    if type_fun == 'group_p1':
        report = GroupP1(user,request)
        report.create()
    elif type_fun == 'group_p2':
        report = GroupP2(user,request)
        report.create()
    elif type_fun == 'group_p3':
        report = GroupP3(user,request)
        report.create()
    elif type_fun == 'group_p4':
        report = GroupP4(user,request)
        report.create()
    elif type_fun == 'group_p5':
        report = GroupP5(user,request)
        report.create()
    elif type_fun == 'group_p6':
        report = GroupP6(user,request)
        report.create()
    elif type_fun == 'group_p7':
        report = GroupP7(user,request)
        report.create()
    elif type_fun == 'group_p8':
        report = GroupP8(user,request)
        report.create()
    elif type_fun == 'group_p9':
        report = GroupP9(user,request)
        report.create()
    elif type_fun == 'group_p10':
        report = GroupP10(user,request)
        report.create()
    elif type_fun == 'group_p11':
        report = GroupP11(user,request)
        report.create()
    elif type_fun == 'group_p12':
        report = GroupP12(user,request)
        report.create()
    elif type_fun == 'group_p13':
        report = GroupP13(user,request)
        report.create()
    elif type_fun == 'group_p14':
        report = GroupP14(user,request)
        report.create()
    elif type_fun == 'group_p15':
        report = GroupP15(user,request)
        report.create()
    elif type_fun == 'group_p16':
        report = GroupP16(user,request)
        report.create()
    elif type_fun == 'group_p17':
        report = GroupP17(user,request)
        report.create()
    elif type_fun == 'group_p18':
        report = GroupP18(user,request)
        report.create()
    elif type_fun == 'group_p19':
        report = GroupP19(user,request)
        report.create()
    elif type_fun == 'group_p20':
        report = GroupP20(user,request)
        report.create()
    elif type_fun == 'group_p21':
        report = GroupP21(user,request)
        report.create()
    elif type_fun == 'group_p22':
        report = GroupP22(user,request)
        report.create()
    elif type_fun == 'group_p23':
        report = GroupP23(user,request)
        report.create()
    elif type_fun == 'group_p24':
        report = GroupP24(user,request)
        report.create()
    elif type_fun == 'group_p25':
        report = GroupP25(user,request)
        report.create()

    if type_fun == 'vault_otd_rep_ttt':
        report = VaultOtdTTT(user,request)
        report.create()
    elif type_fun == 'vault_otd_rep_nnn':
        report = VaultOtdNNN(user,request)
        report.create()
    elif type_fun == 'vault_otd_rep_hh':
        report = VaultOtdHH(user,request)
        report.create()


    type_fun = request.get('type_report')
    if type_fun == 'a_oth_1':
        report = AOth1(user,request)
        report.create()
    elif type_fun == 'a_oth_2':
        report = AOth2(user,request)
        report.create()
    elif type_fun == 'a_oth_3':
        report = AOth3(user,request)
        report.create()
    elif type_fun == 'a_oth_4':
        report = AOth4(user,request)
        report.create()
    elif type_fun == 'a_oth_5':
        report = AOth5(user,request)
        report.create()
    elif type_fun == 'a_oth_6':
        report = AOth6(user,request)
        report.create()
    elif type_fun == 'a_oth_7':
        report = AOth7(user,request)
        report.create()
    elif type_fun == 'a_oth_8':
        report = AOth8(user,request)
        report.create()
    elif type_fun == 'a_oth_9':
        report = AOth9(user,request)
        report.create()
    elif type_fun == 'a_oth_10':
        report = AOth10(user,request)
        report.create()
    elif type_fun == 'a_oth_11':
        report = AOth11(user,request)
        report.create()
    elif type_fun == 'a_oth_12':
        report = AOth12(user,request)
        report.create()
    elif type_fun == 'a_oth_13':
        report = AOth13(user,request)
        report.create()
    elif type_fun == 'a_oth_14':
        report = AOth14(user,request)
        report.create()
    elif type_fun == 'a_oth_15':
        report = AOth15(user,request)
        report.create()
    elif type_fun == 'a_oth_16':
        report = AOth16(user,request)
        report.create()
    elif type_fun == 'a_oth_17':
        report = AOth17(user,request)
        report.create()
    elif type_fun == 'a_oth_18':
        report = AOth18(user,request)
        report.create()
    elif type_fun == 'a_oth_19':
        report = AOth19(user,request)
        report.create()
    elif type_fun == 'a_oth_20':
        report = AOth20(user,request)
        report.create()
    elif type_fun == 'a_oth_21':
        report = AOth21(user,request)
        report.create()
    elif type_fun == 'a_oth_22':
        report = AOth22(user,request)
        report.create()
    elif type_fun == 'a_oth_23':
        report = AOth23(user,request)
        report.create()
    elif type_fun == 'a_oth_24':
        report = AOth24(user,request)
        report.create()
    elif type_fun == 'a_oth_25':
        report = AOth25(user,request)
        report.create()
    elif type_fun == 'a_oth_26':
        report = AOth26(user,request)
        report.create()
    elif type_fun == 'a_oth_27':
        report = AOth27(user,request)
        report.create()
    elif type_fun == 'a_oth_28':
        report = AOth28(user,request)
        report.create()
    elif type_fun == 'a_oth_29':
        report = AOth29(user,request)
        report.create()
    elif type_fun == 'a_oth_30':
        report = AOth30(user,request)
        report.create()
    elif type_fun == 'a_oth_31':
        report = AOth31(user,request)
        report.create()
    elif type_fun == 'a_oth_32':
        report = AOth32(user,request)
        report.create()
    elif type_fun == 'a_oth_33':
        report = AOth33(user,request)
        report.create()
